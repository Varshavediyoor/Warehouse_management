from django.utils import timezone
from django.shortcuts import render

from .forms import UserRegistrationForm
from .models import UserProfile
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from .models import Notification
from django.utils import timezone
from .utils import notify_admins
from django.shortcuts import get_object_or_404
from .models import UserActivityLog
from django.shortcuts import render
from .forms import UserRegistrationForm,UserEditForm,UserProfileEditForm
from .models import UserProfile
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from .models import Notification
from django.utils import timezone
from .models import UserActivityLog
import urllib.parse
from reportlab.lib.pagesizes import letter
#from salesofficer.models import PickListItem 
from django.http import HttpResponse
from reportlab.pdfgen import canvas
#import openpyxl
#from salesofficer.models import PickListItem
#from salesofficer.models import DeliveryAssignment 
from django.db.models import Q
from django.core.paginator import Paginator
from .utils import notify_admins 
from django.utils import timezone
from django.db.models import Q, F, Case, When, Value, CharField
from django.db.models import Q, Prefetch
#from salesofficer.models import PickList
from .forms import AdminProfileEditForm
from .models import RoleFunction, Function, UserProfile
#from maintenance.models import MaintenanceNotification
from .models import Carrier
from .forms import CarrierForm
#from salesofficer.models import Packing
#from admin_app.utils import save_report
from .models import ReportFile
from maintenance.models import MaintenanceNotification
# admin_app/views.py

#from finance.tax_config import DB_TAX_MAP  # <-- ADD THIS

import pytz






#from procurement_officer.models import PurchaseOrder, PurchaseOrderItem, GoodsReceiptNote, GRNItem,Batch


from django.shortcuts import render, redirect
from warehouse.db_router import set_company as router_set_company



from django.shortcuts import render
from warehouse.db_router import set_company

from django.shortcuts import redirect
from django.contrib import messages
from django.conf import settings



from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from .forms import AdminLoginForm, AdminRegisterForm
from django.views.decorators.csrf import csrf_exempt


VALID_COMPANIES = list(settings.DATABASES.keys())

def index(request):
    company_code = request.session.get("company_code")
    return render(request, "home/home.html", {
        "company_code": company_code or "default",
        "show_modal": getattr(request, "show_company_popup", False),
    })

from warehouse.db_router import set_company as set_company_db

def set_company(request):
    if request.method == "POST":
        company_code = request.POST.get("company_code", "").strip()

        if company_code not in VALID_COMPANIES:
            messages.error(request, "Invalid Company Code")
            return redirect("admin_index")

        request.session["company_code"] = company_code
        request.session["company_timezone"] = settings.COMPANY_TIMEZONES.get(
            company_code, settings.DEFAULT_COMPANY_TIMEZONE
        )
        messages.success(request, f"Company selected: {company_code}")
        return redirect("admin_index")

def company_home(request, company):
    # Set the selected company in the session
    request.session["company_code"] = company
    return render(request, "home/home.html", {"company": company})



# -------------------- Registration --------------------
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from .forms import AdminLoginForm, AdminRegisterForm
from .models import AdminProfile
from .utils import generate_otp, send_otp_email
from django.utils import translation
from django.conf import settings
from django.conf import settings
from django.utils import translation
from django.shortcuts import redirect
from django.conf.locale import LANG_INFO
from urllib.parse import urlparse
import re
#from salesofficer.models import Dispatch
#from salesofficer.models import Packing



def switch_language(request, lang_code):
    """
    Switch between languages with proper i18n URL prefix handling.
    """
    if lang_code not in dict(settings.LANGUAGES):
        lang_code = settings.LANGUAGE_CODE

    # Activate new language in session
    translation.activate(lang_code)
    request.session['django_language'] = lang_code

    # Get the previous page
    next_url = request.META.get('HTTP_REFERER', '/')
    parsed = urlparse(next_url)
    path = parsed.path

    # Replace existing /en/ or /ar/ prefix, or add if missing
    if re.match(r'^/(en|ar)/', path):
        path = re.sub(r'^/(en|ar)/', f'/{lang_code}/', path)
    else:
        # Add prefix to path if missing
        path = f'/{lang_code}{path}' if path != '/' else f'/{lang_code}/'

    # Preserve query string if exists
    if parsed.query:
        path = f'{path}?{parsed.query}'

    return redirect(path)

def add_post(request):
    if request.method == "POST":
        text = request.POST.get("text")
        
        # Translate user input to Arabic
        translated_ar = GoogleTranslator(source='auto', target='ar').translate(text)
        
        # Save original + Arabic
        Post.objects.create(
            original_text=text,
            translated_ar=translated_ar
        )
        return redirect("show_posts")

    return render(request, "add_post.html")

def show_posts(request):
    posts = Post.objects.all()
    return render(request, "show_posts.html", {"posts": posts})


from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from .forms import AdminRegisterForm, AdminAccessKeyForm
from .models import AdminProfile


# Step 1: Ask for Security Key
from django.contrib import messages

def admin_access_key(request):
    if request.method == "POST":
        form = AdminAccessKeyForm(request.POST)
        if form.is_valid():
            access_key = form.cleaned_data["access_key"]

            if access_key == settings.ADMIN_SECURITY_KEY:
                # ✅ Store a session flag to allow registration
                request.session["admin_key_verified"] = True

                # ✅ Clear all old messages (including "Invalid security key")
                list(messages.get_messages(request))  # this empties the message queue

                # ✅ Redirect directly to admin registration page
                return redirect("admin_register")

            else:
                # ❌ Wrong key
                messages.error(request, "Invalid security key.")
    else:
        form = AdminAccessKeyForm()

    return render(request, "admin_app/admin_access_key.html", {"form": form})

# Step 2: Actual Admin Registration (only accessible after key check)
def admin_register(request):
    """Custom admin registration with session-based security check"""

    # Security check
    if not request.session.get("admin_key_verified"):
        return render(request, "admin_app/403.html", status=403)

    # Redirect if already logged in
    if request.user.is_authenticated:
        return redirect('custom_admin_dashboard')

    if request.method == 'POST':
        form = AdminRegisterForm(request.POST, request.FILES)
        if form.is_valid():

            # Extract cleaned data
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            phoneno = form.cleaned_data['phoneno']
            profile_photo = form.cleaned_data.get('profile_photo')

            # Create the user (unique validation already handled in form)
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                is_staff=True
            )

            # Create Admin profile
            AdminProfile.objects.create(
                user=user,
                email=email,
                phoneno=phoneno,
                profile_photo=profile_photo,
            )

            # Clear session key after success
            request.session.pop("admin_key_verified", None)

            messages.success(request, f"Admin '{username}' registered successfully! Please login.")
            return redirect('admin_login')
    else:
        form = AdminRegisterForm()

    return render(request, 'admin_app/admin_register.html',{'form':form})



# -------------------- Admin Login with Email OTP --------------------

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.utils import timezone
from django.http import HttpResponseRedirect
from django.urls import reverse
from datetime import timedelta

from .forms import AdminLoginForm
from .models import AdminProfile
from .utils import generate_otp, send_otp_email

#---------------------Admin Login connection with API--------------------
def admin_login_internal(username, password):
    user = authenticate(username=username, password=password)

    if not user or not user.is_staff:
        return {"status": "error", "message": "Invalid username or password"}

    try:
        profile = AdminProfile.objects.get(user=user)
    except AdminProfile.DoesNotExist:
        return {"status": "error", "message": "Admin profile not found"}

    # OTP validity check
    if profile.last_otp_verified_at:
        if timezone.now() - profile.last_otp_verified_at < timedelta(weeks=2):
            return {
                "status": "success",
                "otp_required": False,
                "user_id": user.id
            }

    # OTP required
    otp = generate_otp()
    profile.otp = otp
    profile.otp_created_at = timezone.now()
    profile.otp_verified = False
    profile.save()

    send_otp_email(profile.email, otp)

    return {
        "status": "success",
        "otp_required": True,
        "user_id": user.id,
        "message": f"OTP sent to {profile.email}"
    }



# -------------------- Admin Login with Email OTP --------------------
def admin_login(request):
    """Step 1: Username + Password -> Send OTP only if last OTP is older than 2 weeks"""
    if request.user.is_authenticated:
        return redirect('custom_admin_dashboard')

    # always create a form instance so `form` exists for GET and POST
    form = AdminLoginForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            # use internal function if available (keeps logic shared with API)
            try:
                result = admin_login_internal(username, password)
            except NameError:
                # fallback if internal not present — run the inline logic
                user = authenticate(request, username=username, password=password)
                if not user or not user.is_staff:
                    result = {"status": "error", "message": "Invalid username or password"}
                else:
                    try:
                        profile = AdminProfile.objects.get(user=user)
                    except AdminProfile.DoesNotExist:
                        result = {"status": "error", "message": "Admin profile not found"}
                    else:
                        if profile.last_otp_verified_at and timezone.now() - profile.last_otp_verified_at < timedelta(weeks=2):
                            result = {"status": "success", "otp_required": False, "user_id": user.id}
                        else:
                            otp = generate_otp()
                            profile.otp = otp
                            profile.otp_created_at = timezone.now()
                            profile.otp_verified = False
                            profile.save()
                            send_otp_email(profile.email, otp)
                            result = {"status": "success", "otp_required": True, "user_id": user.id, "message": f"OTP sent to {profile.email}"}

            if result.get("status") == "error":
                messages.error(request, result.get("message", "Login failed."))
                return redirect("admin_login")

            # OTP NOT needed → Direct login
            if not result.get("otp_required"):
                user = User.objects.get(id=result["user_id"])
                login(request, user)
                return redirect("custom_admin_dashboard")

            # OTP needed → Redirect to OTP page
            request.session['login_user_id'] = result["user_id"]
            messages.info(request, result.get("message", "OTP sent."))
            return redirect("verify_admin_login")

        else:
            messages.error(request, "Invalid username or password.")

    # GET request or invalid form => render page with form (empty or with errors)
    return render(request, 'admin_app/admin_login.html', {'form': form})



def verify_admin_login(request):
    """Step 2: Verify OTP and login"""
    user_id = request.session.get('login_user_id')

    # Check for missing session
    if not user_id:
        messages.error(request, "Session expired. Please login again.")
        return redirect('admin_login')

    try:
        profile = AdminProfile.objects.get(user_id=user_id)
    except AdminProfile.DoesNotExist:
        messages.error(request, "Session expired. Please login again.")
        return redirect('admin_login')

    if request.method == 'POST':
        entered_otp = request.POST.get('otp')

        # Validate OTP expiration (30 seconds)
        expiry_time = profile.otp_created_at + timedelta(seconds=30) if profile.otp_created_at else None
        if not expiry_time or timezone.now() > expiry_time:
            messages.error(request, "OTP expired. Please request a new one.")
            return redirect('resend_admin_otp')

        # Check if OTP already used
        if profile.otp_verified:
            messages.error(request, "This OTP has already been used. Please login again.")
            return redirect('admin_login')

        # Match OTP
        if entered_otp == profile.otp:

            # ⭐ Mark OTP as verified + store timestamp for 2-week rule
            profile.otp_verified = True
            profile.last_otp_verified_at = timezone.now()

            # Clear OTP after successful verification
            profile.otp = None
            profile.save()

            # Login user
            login(request, profile.user)
            request.session.set_expiry(settings.SESSION_COOKIE_AGE)

            # Remove session temp id
            request.session.pop('login_user_id', None)

            messages.success(request, f"Welcome, {profile.user.username}!")
            return redirect('custom_admin_dashboard')

        else:
            messages.error(request, "Invalid OTP. Please try again.")

    return render(request, 'admin_app/verify_login.html')



def resend_admin_otp(request):
    """Resend OTP to admin's registered email"""
    user_id = request.session.get('login_user_id')
    if not user_id:
        messages.error(request, "Session expired. Please login again.")
        return redirect('admin_login')

    try:
        profile = AdminProfile.objects.get(user_id=user_id)
    except AdminProfile.DoesNotExist:
        messages.error(request, "Profile not found.")
        return redirect('admin_login')

    # Invalidate old OTP and generate a new one
    otp = generate_otp()
    profile.otp = otp
    profile.otp_created_at = timezone.now()
    profile.otp_verified = False
    profile.save()

    send_otp_email(profile.email, otp)

    # Reset session timeout for another 30 seconds
    request.session.set_expiry(30)

    messages.success(request, f"New OTP sent to {profile.email}. Please verify within 30 seconds.")
    return HttpResponseRedirect(reverse('verify_admin_login'))

@login_required
def admin_profile_view(request):
    user = request.user
    try:
        profile = user.userprofile
    except UserProfile.DoesNotExist:
        profile = None  # handle case if profile is not created yet

    context = {
        'user': user,
        'profile': profile,
    }
    return render(request, 'admin_app/admin_profile_view.html', context)


# -------------------- Dashboard --------------------
def staff_required(user):
    return user.is_authenticated and user.is_staff


#from maintenance.models import MaintenanceNotification

from maintenance.models import MaintenanceNotification
from django.contrib.auth.decorators import user_passes_test
from admin_app.utils import staff_required  # your existing staff check

from django.utils import timezone

@user_passes_test(staff_required)
def custom_admin_dashboard(request):
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    # Current time in **active timezone**
    current_time = timezone.localtime(timezone.now())

    # Pass current_time to template
    return render(request, 'admin_app/admin_dashboard.html', {
        'maintenance_notifications': maintenance_notifications,
        'current_time': current_time,
    })

# -------------------- Logout --------------------
def admin_logout(request):
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('admin_login')  # ✅ Redirect to home


#------------------------------------------------------------#

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.utils import timezone

from .models import UserProfile, Supplier  # A: include Supplier model import
from .forms import UserRegistrationForm
# from .utils import log_user_activity, notify_admins, get_client_ip  # if already used in your project


# ----------------------------
# REGISTER USER VIEW
# ----------------------------

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.utils import timezone
from .models import UserProfile, Supplier, DeliveryBoy, Notification
from .forms import UserRegistrationForm


# ----------------------------
# USER REGISTRATION VIEW
# ----------------------------
def register_user(request):
    if request.method == "POST":
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            # ✅ Create new User
            user = form.save(commit=False)
            user.set_password(form.cleaned_data["password"])
            user.save()

            # ✅ Extract role
            role = form.cleaned_data["role"].strip().upper()

            # ✅ Create UserProfile for the user
            user_profile = UserProfile.objects.create(
                user=user,
                role=role,
                phone_number=form.cleaned_data.get("phone_number", ""),
                gender=form.cleaned_data.get("gender", "Male"),
                nationality=form.cleaned_data.get("nationality", ""),
            )

            # ✅ Handle Supplier creation
            if role == "SUPPLIER":
                Supplier.objects.create(
                    user=user,
                    address=form.cleaned_data.get("address", ""),
                    contact_person=form.cleaned_data.get("contact_person", ""),
                    gst_no=form.cleaned_data.get("gst_no", ""),
                )

            # ✅ Handle Delivery Boy creation
            elif role == "DELIVERY_BOY" or role == "DELIVERY BOY":
                DeliveryBoy.objects.create(
                    user=user,
                    address=form.cleaned_data.get("address", ""),
                    city=form.cleaned_data.get("city", ""),
                    state=form.cleaned_data.get("state", ""),
                    pincode=form.cleaned_data.get("pincode", ""),
                    vehicle_number=form.cleaned_data.get("vehicle_number", ""),
                    vehicle_type=form.cleaned_data.get("vehicle_type", ""),
                    license_number=form.cleaned_data.get("license_number", ""),
                    assigned_area=form.cleaned_data.get("assigned_area", ""),
                    remarks=form.cleaned_data.get("remarks", ""),
                )

            messages.success(request, f"✅ User '{user.username}' registered successfully!")
            return redirect("register_user")

        else:
            messages.error(request, "❌ Please correct the errors below.")
    else:
        form = UserRegistrationForm()

    return render(request, "admin_app/register_user.html", {"form": form})


# ----------------------------
# ROLE-BASED DASHBOARD REDIRECT
# ----------------------------
failed_attempts = {}
FAILED_THRESHOLD = 3  # Maximum allowed failed attempts


def redirect_to_dashboard(role):
    """
    Normalizes role and redirects user to appropriate dashboard.
    Handles space/underscore/case mismatches automatically.
    """
    if not isinstance(role, str):
        return None

    # Normalize: lowercase → uppercase, spaces → underscores
    role_norm = role.strip().replace(" ", "_").upper()

    dashboard_map = {
        "WAREHOUSE_MANAGER": "warehouse_manager_dashboard",
        "INVENTORY_SUPERVISOR": "inventory_supervisor_dashboard",
        "STOREKEEPER": "storekeeper_dashboard",
        "LOGISTICS": "logistics_dashboard",
        "FINANCE": "finance_dashboard",
        "EXTERNAL": "external_user_dashboard",
        "ADMIN": "admin_index",
        "PROCUREMENT_OFFICER": "procurement_officer_dashboard",
        "SUPPLIER": "supplier_dashboard",
        "INVENTORY_MANAGER": "inventory_manager_dashboard",
        "SALES_OFFICER": "salesofficer_dashboard",
        "DELIVERY_BOY": "delivery_boy_dashboard",
        "MAINTENANCE_STAFF": "maintenance_staff_dashboard",
        "SECURITY": "security_dashboard",
        "QC_INSPECTOR":"quality_control"
    }

    if role_norm in dashboard_map:
        print(f"✅ Redirecting role: {role_norm} → {dashboard_map[role_norm]}")
        return redirect(dashboard_map[role_norm])

    print(f"⚠️ Unrecognized role: {role_norm}")
    return None



from django.contrib.auth import login, logout
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from django.utils import timezone
from django.shortcuts import redirect, render
from django.contrib.auth import get_user_model

from admin_app.models import UserSessionTimer
from .utils import (
    user_login_internal,
    log_user_activity,
    notify_admins,
    get_client_ip
)
from .models import UserProfile

User = get_user_model()

failed_attempts = {}
FAILED_THRESHOLD = 3

def login_view(request):
    """
    Handles user login with role validation, session timer, failed login tracking,
    and timestamps in company's timezone.
    """
    # Auto logout if already logged in
    if request.user.is_authenticated:
        log_user_activity(request.user, request.user.username, "auto_logout", request)
        logout(request)
        messages.info(request, "You have been logged out. Please login again.")

    form = AuthenticationForm(request, data=request.POST or None)
    username = request.POST.get("username", "").strip()
    selected_role = request.POST.get("role", "").strip()

    # Determine company timezone from session or default
    company_code = request.session.get("company_code", "default")
    company_tz_str = settings.COMPANY_TIMEZONES.get(company_code, settings.DEFAULT_COMPANY_TIMEZONE)
    company_tz = pytz.timezone(company_tz_str)


    if request.method == "POST":
        if form.is_valid():
            password = form.cleaned_data.get("password")
            result = user_login_internal(username, password)

            if result["status"] == "error":
                messages.error(request, result["message"])
                log_user_activity(None, username, "failed_login", request)
                return redirect("user_login")

            user = User.objects.get(id=result["user_id"])

            # Role check
            try:
                actual_role = user.userprofile.role
            except UserProfile.DoesNotExist:
                messages.error(request, "No role/profile assigned for this account.")
                logout(request)
                return redirect("user_login")

            if selected_role.replace(" ", "_").upper() != actual_role.replace(" ", "_").upper():
                messages.error(request, "Username does not belong to selected role.")
                return redirect("user_login")

            # ✅ LOGIN
            login(request, user)

            # ✅ SESSION TIMER with company timezone
            now_in_company_tz = timezone.now().astimezone(company_tz)
            UserSessionTimer.objects.update_or_create(
                user=user,
                defaults={"login_time": now_in_company_tz}
            )

            # Reset failed attempts
            failed_attempts[username] = 0
            log_user_activity(user, username, "login", request)

            messages.success(request, f"Welcome back, {user.first_name}!")
            return redirect_to_dashboard(actual_role)

        else:
            # Invalid form (username/password)
            messages.error(request, "Invalid username or password.")
            failed_attempts[username] = failed_attempts.get(username, 0) + 1
            log_user_activity(None, username, "failed_login", request)

            if failed_attempts[username] >= FAILED_THRESHOLD:
                ip = get_client_ip(request)
                now_in_company_tz = timezone.now().astimezone(company_tz)
                notify_admins(
                    f"⚠ User '{username}' failed login {FAILED_THRESHOLD} times\n"
                    f"IP: {ip}\nTime: {now_in_company_tz}"
                )
                failed_attempts[username] = 0

            return redirect("user_login")

    # GET request - show login form
    return render(request, "user/login.html", {"form": form})


def password_expired(request):
    return render(request, 'admin_app/password_expired.html')

# ----------------------------
# HELPER FUNCTIONS
# ----------------------------
def get_client_ip(request):
    """Helper to get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0]
    return request.META.get('REMOTE_ADDR')


# ----------------------------
# ADMIN NOTIFICATIONS
# ----------------------------
@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_notifications(request):

    user_notifications = Notification.objects.filter(
        recipient=request.user
    ).order_by('-created_at')

    # maintenance_notes = MaintenanceNotification.objects.filter(
    #     active=True
    # ).order_by('-created_at')

    # ✅ Mark as read FIRST
    user_notifications.filter(is_read=False).update(is_read=True)

    all_notifications = list(user_notifications) + list(maintenance_notes)
    all_notifications.sort(key=lambda x: x.created_at, reverse=True)

    paginator = Paginator(all_notifications, 10)
    notifications = paginator.get_page(request.GET.get('page'))

    return render(request, 'admin_app/admin_notifications.html', {
        'notifications': notifications,   # ✅ ONLY THIS
    })



# ----------------------------
# DASHBOARD VIEWS
# ----------------------------

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.utils import timezone

from admin_app.models import RoleFunction, UserProfile
from maintenance.models import MaintenanceNotification   # import for notifications


@login_required
def qc_dashboard(request):

    # 🕒 ACTIVITY LOGGING (SESSION TIME)
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)  # default 5 mins

    now = timezone.now().timestamp()

    if last_activity:
        remaining = int(timeout - (now - last_activity))
        remaining = max(remaining, 0)
    else:
        remaining = timeout

    # update last activity on each dashboard load
    request.session["last_activity"] = now

    # 🛠 QC FUNCTIONS (STATIC – SIMPLE & SAFE)
    qc_functions = [
        "View GRN",
        "View Purchase Returns",
        "Open QC Tasks",
        "Enter Inspection Result",
        "Approve Stock",
        "Reject Stock",
        "Move Stock – Internal",
        "Move Stock – Scrap",
        "Move Stock – Vendor",
    ]

    # 🔧 MAINTENANCE NOTIFICATIONS
    notifications = MaintenanceNotification.objects.filter(
        active=True
    ).order_by("-created_at")

    context = {
        "remaining_session_time": remaining,
        "qc_functions": qc_functions,
        "notifications": notifications,
        "dashboard_title": "Quality Control Dashboard",
    }

    return render(request, "user/qc_dashboard.html", context)

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone
from django.conf import settings

from .models import Notification  # Your single Notification model
from django.contrib.auth.models import Group

from .models import Notification  # Your single Notification model
from django.contrib.auth.models import Group
#from warehouse_manager.utils import generate_expiry_notifications, generate_low_stock_notifications


@login_required
def warehouse_manager_dashboard(request):
    # ---------------------------
    # 1️⃣ SESSION TIME CALCULATION
    # ---------------------------
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)  # default 5 min
    now = timezone.now().timestamp()
    # maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        remaining = max(0, int(timeout - (now - last_activity)))
    else:
        remaining = timeout

    # -------------------------------------------
    # 2️⃣ GET ALLOWED FUNCTIONS BASED ON USER ROLE
    # -------------------------------------------
    user_role = request.user.userprofile.role

    try:
        role_functions = RoleFunction.objects.get(role=user_role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []

    # -------------------------------------------
    # 3️⃣ TRIGGER NOTIFICATIONS (UTILS HANDLE LOGIC)
    # -------------------------------------------
    # generate_expiry_notifications(request.user)
    # generate_low_stock_notifications()

    # ---------------------------
    # 4️⃣ FETCH NOTIFICATIONS
    # ---------------------------
    notifications = request.user.notifications.filter(
        is_read=False
    ).order_by("-created_at")

    # ---------------------------
    # 5️⃣ CONTEXT
    # ---------------------------
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,

        # Dashboard metrics (placeholders)
        "total_pos": 0,
        "pos_this_week": 0,
        "pending_grns": 0,
        "suppliers": 0,
        "received_batches": 0,

        # Notifications
        "notifications": notifications,
        # 'maintenance_notifications': maintenance_notifications,
    }

    return render(request, "user/warehouse_manager.html", context)

@login_required
def inventory_supervisor_dashboard(request):

    # ---------------------------
    # 1️⃣ SESSION TIME CALCULATION
    # ---------------------------
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)  # default 5 min
    # maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        if remaining < 0:
            remaining = 0
    else:
        remaining = timeout

    # -------------------------------------------
    # 2️⃣ GET ALLOWED FUNCTIONS BASED ON USER ROLE
    # -------------------------------------------
    user_role = request.user.userprofile.role

    try:
        role_functions = RoleFunction.objects.get(role=user_role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []

    # ---------------------------
    # 3️⃣ CREATE NOTIFICATIONS FOR EXPIRING STOCK
    # ---------------------------
    today = timezone.now().date()
    expiring_soon_days = 7

    batches = Batch.objects.filter(expiry_date__isnull=False)
    for batch in batches:
        days_remaining = (batch.expiry_date - today).days

        if days_remaining < 0:
            message = f"Batch {batch.batch_number} of {batch.stock.name} has expired!"
        elif days_remaining <= expiring_soon_days:
            message = f"Batch {batch.batch_number} of {batch.stock.name} will expire in {days_remaining} days."
        else:
            continue  # skip batches far from expiry

        # Only create notification if it doesn't exist yet
        exists = Notification.objects.filter(
            recipient=request.user,
            message=message,
            is_read=False
        ).exists()
        if not exists:
            Notification.objects.create(
                recipient=request.user,
                message=message
            )

    # ---------------------------
    # 4️⃣ GET NOTIFICATIONS TO DISPLAY
    # ---------------------------
    notifications = request.user.notifications.filter(is_read=False).order_by('-created_at')

    # ---------------------------
    # 5️⃣ FINAL CONTEXT
    # ---------------------------
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,

        # Dashboard counts (adjust as needed)
        "total_pos": 0,
        "pos_this_week": 0,
        "pending_grns": 0,
        "suppliers": 0,
        "received_batches": 0,

        # Notifications
        "notifications": notifications,
        # 'maintenance_notifications': maintenance_notifications
    }

    return render(request, 'user/inventory_supervisor.html', context)

@login_required
def storekeeper_dashboard(request):

    # ---------------------------
    # 1️⃣ SESSION TIME CALCULATION
    # ---------------------------
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)  # 5 minutes
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        if remaining < 0:
            remaining = 0
    else:
        remaining = timeout

    # -------------------------------------------
    # 2️⃣ GET ALLOWED FUNCTIONS BASED ON USER ROLE
    # -------------------------------------------
    user_role = request.user.userprofile.role

    try:
        role_functions = RoleFunction.objects.get(role=user_role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []

    # ---------------------------
    # 3️⃣ FINAL CONTEXT
    # ---------------------------
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,

        # Add real dashboard values later:
        "total_pos": 0,
        "pos_this_week": 0,
        "pending_grns": 0,
        "suppliers": 0,
        "received_batches": 0,
        'maintenance_notifications': maintenance_notifications
    }

    return render(request, 'user/storekeeper.html', context)


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.utils import timezone

from admin_app.models import RoleFunction, UserProfile
#from maintenance.models import MaintenanceNotification   # import for notifications


@login_required
def qc_dashboard(request):

    # 1️⃣ GET SESSION TIME LEFT
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        remaining = max(remaining, 0)
    else:
        remaining = timeout

    # 2️⃣ GET ROLE-BASED PERMISSIONS
    user_role = getattr(request.user.userprofile, "role", None)

    DEFAULT_QC_FUNCTIONS = [
        "View GRN & Returns",
        "Open QC Tasks",
        "Enter inspection result",
        "Approve / Reject stock",
        "Move stock: Internal",
        "Move stock: Scrap",
        "Move stock: Vendor",
        
    ]

    try:
        role_functions = RoleFunction.objects.get(role=user_role)
        allowed_functions = list(role_functions.functions.values_list("name", flat=True))
    except RoleFunction.DoesNotExist:
        allowed_functions = DEFAULT_QC_FUNCTIONS if user_role == "QC_INSPECTOR" else []

    # 3️⃣ NOTIFICATIONS
    notifications = MaintenanceNotification.objects.filter(active=True).order_by("-created_at")

    # 4️⃣ PASS DATA
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,
        "notifications": notifications,
        "dashboard_title": "Quality Control Dashboard",
        "role_code": "QC_INSPECTOR",
        'maintenance_notifications': maintenance_notifications,
    }

    return render(request, "user/qc_dashboard.html", context)








from datetime import datetime, time, timedelta

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.shortcuts import render
from django.utils import timezone

#from salesofficer.models import SalesOrder,DeliveryOrder,ReturnRequest
#from finance.models import SalesInvoice
#from maintenance.models import MaintenanceNotification
from admin_app.models import RoleFunction


from datetime import datetime, timedelta, time

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.shortcuts import render
from django.utils import timezone




from datetime import datetime, timedelta, time

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.shortcuts import render
from django.utils import timezone




@login_required
def salesofficer_dashboard(request):
    # Simply render the template
    return render(request, "user/salesofficer.html")


@login_required
def logistics_dashboard(request):

    # ---------------------------
    # 1️⃣ SESSION TIME CALCULATION
    # ---------------------------
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        if remaining < 0:
            remaining = 0
    else:
        remaining = timeout

    # -------------------------------------------
    # 2️⃣ ALLOWED FUNCTIONS FOR USER ROLE
    # -------------------------------------------
    user_role = request.user.userprofile.role

    try:
        role_functions = RoleFunction.objects.get(role=user_role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []

    # -------------------------------------------
    # 3️⃣ CONTEXT
    # -------------------------------------------
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,
        'maintenance_notifications': maintenance_notifications,

        "total_pos": 0,
        "pos_this_week": 0,
        "pending_grns": 0,
        "suppliers": 0,
        "received_batches": 0,
    }

    return render(request, 'user/logistics.html', context)


#from finance.models import SalesInvoice
#from procurement_officer.models import GRNItem
from django.db.models import Sum, F
from decimal import Decimal
#from finance.services import get_finance_metrics

@login_required
def finance_dashboard(request):

    # ---------------------------
    # 1️⃣ SESSION TIME CALCULATION
    # ---------------------------
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        if remaining < 0:
            remaining = 0
    else:
        remaining = timeout

    # -------------------------------------------
    # 2️⃣ ALLOWED FUNCTIONS FOR USER ROLE
    # -------------------------------------------
    user_role = request.user.userprofile.role

    try:
        role_functions = RoleFunction.objects.get(role=user_role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []
    current_db = request.session.get('db_name', 'INDIA001')  # fallback to INDIA001

    tax_context = DB_TAX_MAP.get(current_db, {'registration_url': '#', 'tax_code': 'N/A'})
    


    # =========================
    # REAL REVENUE
    # =========================
    revenue_month = (
        SalesInvoice.objects.aggregate(total=Sum("subtotal"))["total"]
        or Decimal("0.00")
    )

    revenue_today = revenue_month  # temporary (no date split yet)

    # =========================
    # DUMMY OPERATIONAL COSTS
    # =========================
    expenses_today = Decimal("12800.00")
    expenses_month = Decimal("215000.00")

    # =========================
    # PROFIT
    # =========================
    net_profit = revenue_month - expenses_month
    profit_margin = round((net_profit / revenue_month) * 100, 2) if revenue_month else 0

    # =========================
    # INVENTORY VALUATION (REAL)
    # =========================
    inventory_value = (
        GRNItem.objects.filter(
            grn__approval_status="Approved",
            verification_status="Approved"
        ).aggregate(
            total=Sum(F("quantity_received") * F("price"))
        )["total"]
        or Decimal("0.00")
    )
    

    metrics = get_finance_metrics()

    context = {
        **metrics,

        # TEMP until next phase
        "inventory_holding_cost": Decimal("0.00"),
        "fulfillment_cost": Decimal("0.00"),
        "cost_per_order": Decimal("0.00"),

        # CHART DATA (backend driven)
        "chart_labels": ["Jan", "Feb", "Mar", "Apr", "May", "Jun"],
        "chart_revenue": [],
        "chart_expenses": [],
        "chart_profit": [],
        "chart_inventory": [],
        'tax_registration_url': tax_context['registration_url'],
        'tax_code': tax_context['tax_code'],
    }

    # context = {
    #     "revenue_today": revenue_today,
    #     "revenue_month": revenue_month,
    #     "expenses_today": expenses_today,
    #     "expenses_month": expenses_month,
    #     "outstanding_amount": Decimal("0.00"),
    #     "overdue_invoices": SalesInvoice.objects.filter(status="Pending").count(),

    #     "net_profit": net_profit,
    #     "profit_margin": profit_margin,
    #     "inventory_value": inventory_value,
    #     "inventory_holding_cost": Decimal("32000.00"),
    #     "fulfillment_cost": Decimal("18500.00"),
    #     "return_damage_cost": Decimal("9200.00"),
    #     "cost_per_order": Decimal("2500.00"),

    #     # Graph data
    #     "chart_labels": ["Jan", "Feb", "Mar", "Apr", "May", "Jun"],
    #     "chart_revenue": [12000, 18000, 25000, 30000, 35000, float(revenue_month)],
    #     "chart_expenses": [10000, 14000, 17000, 20000, 22000, float(expenses_month)],
    # }

    return render(request, 'user/finance.html', context)

@login_required
def maintenance_dashboard(request):
    # ---------------------------
    # 1️⃣ SESSION TIME CALCULATION
    # ---------------------------
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        if remaining < 0:
            remaining = 0
    else:
        remaining = timeout

    # -------------------------------------------
    # 2️⃣ GET ALLOWED FUNCTIONS BASED ON ROLE
    # -------------------------------------------
    user_role = getattr(request.user.userprofile, "role", None)

    try:
        role_functions = RoleFunction.objects.get(role=user_role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []

    # ---------------------------
    # 3️⃣ FETCH ACTIVE MAINTENANCE NOTIFICATIONS
    # ---------------------------
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True).order_by('-created_at')

    # ---------------------------
    # 4️⃣ CONTEXT
    # ---------------------------
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,
        "notifications": maintenance_notifications,  # You can keep both if needed
        "dashboard_title": "Maintenance Staff Dashboard",
        'maintenance_notifications': maintenance_notifications,  # ✅ fixed
    }

    return render(request, 'user/maintenance_dashboard.html', context)


def security_dashboard(request):
    # Get all dispatches, newest first
    # dispatches = Dispatch.objects.prefetch_related('orders').select_related('dispatched_by').order_by('-dispatch_date')
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    context = {
        'dashboard_title': 'Security Dashboard',
        'remaining_session_time': 3600,  # example
        # 'dispatches': dispatches,
        'maintenance_notifications': maintenance_notifications
    }
    return render(request, 'user/security_dashboard.html', context)



@login_required
def external_user_dashboard(request):

    # ---------------------------
    # 1️⃣ SESSION TIME
    # ---------------------------
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        if remaining < 0:
            remaining = 0
    else:
        remaining = timeout

    # ---------------------------
    # 2️⃣ ALLOWED FUNCTIONS
    # ---------------------------
    role = request.user.userprofile.role
    try:
        role_functions = RoleFunction.objects.get(role=role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []

    # ---------------------------
    # 3️⃣ CONTEXT
    # ---------------------------
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,

        "total_pos": 0,
        "pos_this_week": 0,
        "pending_grns": 0,
        "suppliers": 0,
        "received_batches": 0,
        'maintenance_notifications': maintenance_notifications
    }

    return render(request, 'user/external_user.html', context)

@login_required
def procurement_officer_dashboard(request):

    # 1️⃣ SESSION TIME
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        if remaining < 0:
            remaining = 0
    else:
        remaining = timeout

    password_minutes_left = None

    try:
        timer = UserSessionTimer.objects.get(user=request.user)
        password_minutes_left = timer.minutes_left()
    except UserSessionTimer.DoesNotExist:
        pass


    

    # 2️⃣ ALLOWED FUNCTIONS
    role = request.user.userprofile.role
    try:
        role_functions = RoleFunction.objects.get(role=role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []

    # 3️⃣ CONTEXT
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,

        "total_pos": 0,
        "pos_this_week": 0,
        "pending_grns": 0,
        "suppliers": 0,
        "received_batches": 0,
        'maintenance_notifications': maintenance_notifications
    }

    return render(request, 'user/procurement_officer.html', context)

@login_required
def inventory_manager_dashboard(request):

    # 1️⃣ SESSION TIME
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        if remaining < 0:
            remaining = 0
    else:
        remaining = timeout

    # 2️⃣ ALLOWED FUNCTIONS
    role = request.user.userprofile.role
    try:
        role_functions = RoleFunction.objects.get(role=role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []

    # 3️⃣ NOTIFICATIONS
    # Mark all unread notifications as read
    request.user.notifications.filter(is_read=False).update(is_read=True)
    
    # Fetch latest 5 notifications for display
    notifications = request.user.notifications.order_by('-created_at')[:5]

    # 4️⃣ CONTEXT
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,

        "total_pos": 0,
        "pos_this_week": 0,
        "pending_grns": 0,
        "suppliers": 0,
        "received_batches": 0,

        "notifications": notifications,  # pass notifications to template
        'maintenance_notifications': maintenance_notifications
    }

    return render(request, 'user/inventory_manager.html', context)


@login_required
def supplier_dashboard(request):
    return render(request, 'user/supplier_dashboard.html')

@login_required
def delivery_boy_dashboard(request):

    # ---------------------------
    # 1️⃣ SESSION TIME CALCULATION
    # ---------------------------
    last_activity = request.session.get("last_activity")
    timeout = getattr(settings, "SESSION_COOKIE_AGE", 300)  # 5 minutes
    maintenance_notifications = MaintenanceNotification.objects.filter(active=True)

    if last_activity:
        now = timezone.now().timestamp()
        remaining = int(timeout - (now - last_activity))
        if remaining < 0:
            remaining = 0
    else:
        remaining = timeout

    # -------------------------------------------
    # 2️⃣ GET ALLOWED FUNCTIONS BASED ON USER ROLE
    # -------------------------------------------
    user_role = request.user.userprofile.role

    try:
        role_functions = RoleFunction.objects.get(role=user_role)
        allowed_functions = role_functions.functions.values_list("name", flat=True)
    except RoleFunction.DoesNotExist:
        allowed_functions = []

    # ---------------------------
    # 3️⃣ CONTEXT DATA
    # ---------------------------
    context = {
        "remaining_session_time": remaining,
        "allowed_functions": allowed_functions,
        'maintenance_notifications': maintenance_notifications,

        "total_pos": 0,
        "pos_this_week": 0,
        "pending_grns": 0,
        "suppliers": 0,
        "received_batches": 0,
    }

    return render(request, 'user/delivery_dashboard.html', context)


# ---------------- Logout ----------------
def logout_view(request):
    if request.user.is_authenticated:
        # Log the logout activity
        log_user_activity(request.user, request.user.username, 'logout', request)
        
    logout(request)
    messages.success(request, "Logged out successfully.")
    return redirect('user_login')

def admin_logs(request):
    if not request.user.is_staff:  # optional: restrict access
        messages.error(request, "Access denied.")
        return redirect('admin_login')

    logs = UserActivityLog.objects.all().order_by('-timestamp')
    return render(request, 'admin_app/admin_logs.html', {'logs': logs})

# 
from admin_app.utils import staff_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
from admin_app.models import UserProfile
from admin_app.forms import UserRegistrationForm
from admin_app.utils import log_user_activity, staff_required


# ✅ List all users
@login_required
@user_passes_test(staff_required)
def admin_user_list(request):
    """Displays list of all users for admin."""
    users = User.objects.all().select_related('userprofile')
    return render(request, 'admin_app/admin_user_list.html', {'users': users})

@login_required
def admin_profile_edit(request):
    admin_profile, created = AdminProfile.objects.get_or_create(user=request.user)

    if request.method == "POST":
        form = AdminProfileEditForm(
            request.POST,
            request.FILES,
            instance=admin_profile,
            user=request.user
        )
        if form.is_valid():
            # Save AdminProfile fields
            form.save()

            # Update Django User model fields
            request.user.first_name = form.cleaned_data['first_name']
            request.user.last_name = form.cleaned_data['last_name']
            request.user.email = form.cleaned_data['email']  # use AdminProfile email
            request.user.save()

            return redirect('admin_profile_view')
    else:
        form = AdminProfileEditForm(instance=admin_profile, user=request.user)

    return render(request, "admin_app/admin_profile_edit.html", {"form": form})


@login_required
@user_passes_test(staff_required)
def edit_user(request, user_id):
    user_obj = get_object_or_404(User, id=user_id)
    profile, _ = UserProfile.objects.get_or_create(user=user_obj)

    if request.method == "POST":
        form = UserRegistrationForm(
            request.POST,
            request.FILES,
            instance=user_obj
        )

        if form.is_valid():
            user = form.save(commit=False)

            # 🔑 Update password ONLY if admin entered one
            password = form.cleaned_data.get("password")
            if password:
                user.set_password(password)

            user.save()

            # 🔐 Keep session valid (important!)
            if password and request.user.pk == user.pk:
                update_session_auth_hash(request, user)

            # ✅ Update profile fields
            profile.role = form.cleaned_data.get("role")
            profile.phone_number = form.cleaned_data.get("phone_number")
            profile.gender = form.cleaned_data.get("gender")
            profile.nationality = form.cleaned_data.get("nationality")

            if form.cleaned_data.get("profile_photo"):
                profile.profile_photo = form.cleaned_data.get("profile_photo")

            profile.save()

            messages.success(
                request,
                f"✅ User '{user.username}' updated successfully!"
            )
            return redirect("admin_user_list")

        else:
            messages.error(request, "❌ Please correct the errors below.")

    else:
        # 🔥 Prefill profile fields
        form = UserRegistrationForm(
            instance=user_obj,
            initial={
                "role": profile.role,
                "phone_number": profile.phone_number,
                "gender": profile.gender,
                "nationality": profile.nationality,
            }
        )

    return render(request, "admin_app/register_user.html", {
        "form": form,
        "is_edit": True,
        "edit_user_obj": user_obj,
    })



from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages

from .forms import SimplePasswordResetForm

# from admin_app.models import UserSecurity
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required

@login_required
def reset_password(request):
    if request.method == "POST":
        form = SimplePasswordResetForm(request.POST)
        if form.is_valid():
            user = request.user
            user.set_password(form.cleaned_data["new_password"])
            user.save()

            # security, _ = UserSecurity.objects.get_or_create(user=user)
            # security.password_reset_at = timezone.now()
            # security.save()

            # clear warning flag
            request.session.pop("password_warning_shown", None)
            messages.success(
                request,
                "Password reset successfully. Please log in again."
            )
            return redirect("user_login")
    else:
        form = SimplePasswordResetForm()

    return render(
        request,
        "admin_app/reset_password_confirm.html",
        {"form": form},
    )






@login_required
@user_passes_test(staff_required)
def delete_user(request, user_id):
    user_obj = get_object_or_404(User, id=user_id)
    username = user_obj.username

    if user_obj.is_superuser:
        messages.error(request, "❌ Cannot delete superadmin accounts.")
        return redirect('admin_user_list')

    if request.user == user_obj:
        messages.error(request, "❌ You cannot delete your own account.")
        return redirect('admin_user_list')

    if request.method == "POST":
        user_obj.delete()
        log_user_activity(request.user, username, 'delete_user', request)
        messages.success(request, f"✅ User '{username}' deleted successfully.")
    
    return redirect('admin_user_list')  # always redirect


from django.shortcuts import get_object_or_404

def log_user_activity(user, username, action, request=None):
    """
    Logs user activity for login, logout, failed login
    - user: User instance (can be None for failed login)
    - username: string
    - action: 'login', 'logout', 'failed_login'
    - request: optional, used to get IP
    """
    ip = get_client_ip(request) if request else None

    UserActivityLog.objects.create(
        user=user,
        username=username,
        action=action,
        ip_address=ip,
        timestamp=timezone.now()
    )  # ✅ properly closed



#---------------------- user notification -------------------------


@login_required
def user_profile(request):
    """
    Render the profile page for the logged-in storekeeper.
    """
    return render(request, "user/user_profile.html")

@login_required
def user_notification(request):
    # Fetch all notifications
    notifications = request.user.notifications.order_by('-created_at')
    
    # Count unread notifications
    unread_count = notifications.filter(is_read=False).count()
    
    # Optionally mark as read if opening this page
    notifications.filter(is_read=False).update(is_read=True)

    context = {
        'notifications': notifications,
        'unread_count': unread_count,  # pass unread count
    }
    
    return render(request, "user/user_notification.html", context)


@login_required
def user_delete_notification(request, notification_id):
    notification = Notification.objects.filter(id=notification_id, recipient=request.user).first()
    if notification:
        notification.delete()
    # Redirect back to notifications page anyway
    return redirect('user_notifications')

@login_required
def user_delete_all_notifications(request):
    if request.method == "POST":
        request.user.notifications.all().delete()
    return redirect("user_notifications")


# --- Password reset with admin approval views ---


# admin_app/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test
from django.utils import timezone
from django.core.mail import send_mail
from django.urls import reverse
from django.http import HttpResponseRedirect

from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes

from .forms import ForgotPasswordRequestForm, AdminApproveResetForm, SetNewPasswordForm
from .models import PasswordResetRequest

def forgot_password_request(request):
    if request.method == 'POST':
        form = ForgotPasswordRequestForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            reason = form.cleaned_data['reason']

            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                messages.error(request, 'No user with that email found.')
                return redirect('forgot_password')

            # Create password reset request entry
            PasswordResetRequest.objects.create(
                user=user,
                email=email,
                reason=reason
            )

            messages.success(
                request,
                'Password reset request submitted.'
            )
            # redirect to a page of your choice; e.g. login page
            return redirect('forgot_password')
    else:
        form = ForgotPasswordRequestForm()

    return render(request, 'admin_app/forgot_password.html', {'form': form})

@user_passes_test(lambda u: u.is_staff)
def admin_reset_requests(request):
    # list pending requests (not processed)
    pending = PasswordResetRequest.objects.filter(processed=False).order_by('-requested_at')
    return render(request, 'admin_app/admin_reset_requests.html', {'pending': pending})

@user_passes_test(lambda u: u.is_staff)
def admin_notifications(request):
    notifications = request.user.notifications.order_by('-created_at')
    return render(request, 'admin_app/admin_notifications.html', {
        'notifications': notifications
    })


@user_passes_test(lambda u: u.is_staff)
def approve_reset_request(request, request_id):
    pr = get_object_or_404(PasswordResetRequest, id=request_id)

    if request.method == 'POST':
        form = AdminApproveResetForm(request.POST)
        if form.is_valid() and form.cleaned_data.get('approve'):
            # Mark approved and processed
            pr.approved = True
            pr.approved_by = request.user
            pr.approved_at = timezone.now()
            pr.processed = True
            pr.save()

            # Generate token and uid
            token = default_token_generator.make_token(pr.user)
            uid = urlsafe_base64_encode(force_bytes(pr.user.pk))

            # Build reset URL
            reset_path = reverse('custom_password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
            reset_url = request.build_absolute_uri(reset_path)

            # Email - you can create a nicer HTML template later.
            subject = 'Password reset link'
            message = (
                f'You (or someone on your behalf) requested a password reset.\n\n'
                f'Use the link below to reset your password (this link is single-use / time-limited by Django token):\n\n'
                f'{reset_url}\n\n'
                'If you did not request this, ignore this email.'
            )
            # `from_email` should be set in settings or pass here.
            send_mail(subject, message, None, [pr.email])

            messages.success(request, f'Password reset link sent to {pr.email}')
            return redirect('admin_reset_requests')
        else:
            messages.info(request, 'Approval not confirmed.')
            return redirect('admin_reset_requests')
    else:
        form = AdminApproveResetForm()

    return render(request, 'admin_app/approve_reset.html', {'pr': pr, 'form': form})

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.models import User
from django.utils.http import urlsafe_base64_decode
from django.contrib.auth.tokens import default_token_generator
from django.utils import timezone

from .forms import SetNewPasswordForm
from admin_app.models import UserSessionTimer


def custom_password_reset_confirm(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is None:
        messages.error(request, "Invalid reset link.")
        return redirect('forgot_password')

    # Validate token
    if not default_token_generator.check_token(user, token):
        messages.error(request, "Token expired or invalid.")
        return redirect('forgot_password')

    if request.method == 'POST':
        form = SetNewPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            user.set_password(new_password)
            user.save()

            # ✅ RESET 30-MIN SESSION TIMER AFTER PASSWORD RESET
            UserSessionTimer.objects.update_or_create(
                user=user,
                defaults={'login_time': timezone.now()}
            )

            messages.success(
                request,
                "Password updated successfully. Please login."
            )
            return redirect('user_login')

    else:
        form = SetNewPasswordForm()

    return render(
        request,
        'admin_app/reset_password_confirm.html',
        {'form': form}
    )



from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.urls import reverse
from django.core.mail import send_mail
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.utils import timezone

@login_required
def admin_reports(request):
    return render(request, 'admin_app/admin_reports.html')

@user_passes_test(lambda u: u.is_staff)
def approve_reset_request(request, request_id):
    pr = get_object_or_404(PasswordResetRequest, id=request_id)

    if request.method == 'POST':
        form = AdminApproveResetForm(request.POST)
        if form.is_valid() and form.cleaned_data.get('approve'):

            # ✅ Update request
            pr.approved = True
            pr.approved_by = request.user
            pr.approved_at = timezone.now()
            pr.processed = True
            pr.save()

            # 🔐 Generate reset token
            token = default_token_generator.make_token(pr.user)
            uid = urlsafe_base64_encode(force_bytes(pr.user.pk))

            reset_path = reverse(
                'custom_password_reset_confirm',
                kwargs={'uidb64': uid, 'token': token}
            )
            reset_url = request.build_absolute_uri(reset_path)

            # 📧 Send reset link to staff
            send_mail(
                subject='Password Reset Approved',
                message=(
                    f'Hello {pr.user.username},\n\n'
                    f'Your password reset request has been approved.\n\n'
                    f'Click below to reset your password:\n{reset_url}\n\n'
                    f'This link is valid for a limited time.'
                ),
                from_email=None,
                recipient_list=[pr.email],
            )

            # 🔔 ADMIN DASHBOARD NOTIFICATION (USING YOUR MODEL)
            Notification.objects.create(
                recipient=request.user,   # admin who approved
                message=f"Password reset approved for {pr.user.username}"
            )

            messages.success(request, 'Password reset approved and email sent.')
            return redirect('admin_reset_requests')

    else:
        form = AdminApproveResetForm()

    return render(request, 'admin_app/approve_reset.html', {'pr': pr, 'form': form})


# @user_passes_test(lambda u: u.is_staff)
# def approve_reset_request(request, request_id):
#     pr = get_object_or_404(PasswordResetRequest, id=request_id)
    
#     if request.method == 'POST':
#         form = AdminApproveResetForm(request.POST)
#         if form.is_valid() and form.cleaned_data.get('approve'):
#             # Mark approved and processed
#             pr.approved = True
#             pr.approved_by = request.user
#             pr.approved_at = timezone.now()
#             pr.processed = True
#             pr.save()

#             # Generate token and uid
#             token = default_token_generator.make_token(pr.user)
#             uid = urlsafe_base64_encode(force_bytes(pr.user.pk))

#             # Use your actual URL name here
#             reset_path = reverse('custom_password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
#             reset_url = request.build_absolute_uri(reset_path)

#             # Send email
#             subject = 'Password reset link'
#             message = f'Your password reset link: {reset_url}\nThis link will allow you to reset your password.'
#             send_mail(subject, message, None, [pr.email])

#             messages.success(request, f'Password reset link sent to {pr.email}')
#             return redirect('admin_reset_requests')
#         else:
#             messages.info(request, 'Approval not confirmed.')
#             return redirect('admin_reset_requests')
#     else:
#         form = AdminApproveResetForm()
    
#     return render(request, 'admin_app/approve_reset.html', {'pr': pr, 'form': form})

from django.db.models import Q
  # adjust if supplier model is elsewhere

def purchase_report(request):
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '')
    supplier_filter = request.GET.get('supplier', '')

    # Base Query
    report_data = PurchaseOrderItem.objects.select_related('po', 'po__supplier')

    # Filters
    if search_query:
        report_data = report_data.filter(
            Q(product_name__icontains=search_query)
            | Q(category__icontains=search_query)
            | Q(po__supplier__supplier_name__icontains=search_query)
            | Q(po__po_number__icontains=search_query)
        )
    if category_filter:
        report_data = report_data.filter(category=category_filter)
    if supplier_filter:
        report_data = report_data.filter(po__supplier__supplier_name=supplier_filter)

    final_report = []
    for item in report_data:
        # Qty received from GRN items linked to this PO
        grn_items = GRNItem.objects.filter(
            grn__purchase_order=item.po,
            product_name=item.product_name
        )
        qty_received = grn_items.aggregate(total=Sum('quantity_received'))['total'] or 0
        batch_numbers = grn_items.values_list('batch_no', flat=True).distinct()

        # Determine status
        if qty_received == 0:
            status = "Pending"
        elif qty_received < item.quantity_ordered:
            status = "Partially Received"
        else:
            status = "Completed"

        final_report.append({
            'po_number': item.po.po_number,
            'supplier_name': item.po.supplier.supplier_name or item.po.supplier.user.username,
            'product_name': item.product_name,
            'category': item.category,
            'quantity_ordered': item.quantity_ordered,
            'quantity_received': qty_received,
            'unit_price': 0,  # placeholder
            'total_amount': 0, # placeholder
            'received_date': item.po.order_date,
            'batch_number': ', '.join(batch_numbers) if batch_numbers else '-',
            'status': status
        })

    # Filters for dropdown
    categories = PurchaseOrderItem.objects.values_list('category', flat=True).distinct().order_by('category')
    suppliers = Supplier.objects.annotate(
        display_name=Case(
            When(supplier_name__exact='', then=F('user__username')),
            default=F('supplier_name'),
            output_field=CharField()
        )
    ).values_list('display_name', flat=True).distinct().order_by('display_name')

    context = {
        'report_data': final_report,
        'search_query': search_query,
        'category_filter': category_filter,
        'supplier_filter': supplier_filter,
        'categories': categories,
        'suppliers': suppliers,
        'generated_on': timezone.now(),
    }

    return render(request, 'reports/purchase_report.html', context)

from django.db.models import Sum, Q
from django.utils import timezone


def get_purchase_report_data(request):
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '')
    supplier_filter = request.GET.get('supplier', '')

    report_data = PurchaseOrderItem.objects.select_related(
        'po', 'po__supplier'
    )

    if search_query:
        report_data = report_data.filter(
            Q(product_name__icontains=search_query)
            | Q(category__icontains=search_query)
            | Q(po__supplier__supplier_name__icontains=search_query)
            | Q(po__po_number__icontains=search_query)
        )

    if category_filter:
        report_data = report_data.filter(category=category_filter)

    if supplier_filter:
        report_data = report_data.filter(po__supplier__supplier_name=supplier_filter)

    final_report = []

    for item in report_data:
        grn_items = GRNItem.objects.filter(
            grn__purchase_order=item.po,
            product_name=item.product_name
        )

        qty_received = grn_items.aggregate(
            total=Sum('quantity_received')
        )['total'] or 0

        batch_numbers = ", ".join(
            grn_items.values_list('batch_no', flat=True).distinct()
        ) or "-"

        if qty_received == 0:
            status = "Pending"
        elif qty_received < item.quantity_ordered:
            status = "Partially Received"
        else:
            status = "Completed"

        unit_price = getattr(item, "unit_price", 0)
        total_amount = unit_price * item.quantity_ordered

        final_report.append({
            "po_number": item.po.po_number,
            "supplier": item.po.supplier.supplier_name or item.po.supplier.user.username,
            "product": item.product_name,
            "category": item.category,
            "qty_ordered": item.quantity_ordered,
            "qty_received": qty_received,
            "unit_price": unit_price,
            "total": total_amount,
            "received_date": item.po.order_date,
            "batch": batch_numbers,
            "status": status,
        })

    return {
        "report_data": final_report,
        "generated_on": timezone.now()
    }

    
import os
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from django.utils import timezone
from django.contrib.auth.decorators import login_required

@login_required
def purchase_report_excel(request):
    context = get_purchase_report_data(request)
    
    # Render Excel content
    response = render(request, "reports/purchase_report_export.html", context)
    response["Content-Type"] = "application/vnd.ms-excel"

    # Filename
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Purchase_Report_{timestamp}.xls"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Save to media folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        f.write(response.content)

    # Log in database
    ReportFile.objects.create(
        report_type="Purchase",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


@login_required
def purchase_report_pdf(request):
    context = get_purchase_report_data(request)

    # Render PDF content
    html_string = render_to_string("reports/purchase_report_export.html", context)
    pdf = HTML(string=html_string).write_pdf()

    # Filename
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Purchase_Report_{timestamp}.pdf"

    # Save to media folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        f.write(pdf)

    # Log in database
    ReportFile.objects.create(
        report_type="Purchase",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    # Return PDF as download
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response



def batch_report(request):
    search_query = request.GET.get('search', '').strip()
    category_filter = request.GET.get('category', '')
    supplier_filter = request.GET.get('supplier', '')

    # --- Base Query: fetch batches with related stock and supplier ---
    batches = Batch.objects.select_related('stock', 'stock__supplier').all()

    # --- Apply search filter ---
    if search_query:
        batches = batches.filter(
            Q(batch_number__icontains=search_query) |
            Q(stock__name__icontains=search_query) |
            Q(stock__category__icontains=search_query) |
            Q(stock__supplier__supplier_name__icontains=search_query)
        )

    # --- Apply category filter ---
    if category_filter:
        batches = batches.filter(stock__category=category_filter)

    # --- Apply supplier filter ---
    if supplier_filter:
        batches = batches.filter(stock__supplier__supplier_name=supplier_filter)

    # --- Prepare data for template ---
    batch_data = []
    for b in batches:
        stock = b.stock

        # Total quantity received for this batch from GRN items
        received_total = GRNItem.objects.filter(
            batch_no=b.batch_number
        ).aggregate(total_received=Sum('quantity_received'))['total_received'] or 0

        # Determine status
        status = 'Pending' if received_total < b.quantity_available else 'Packed'

        batch_data.append({
            'batch_number': b.batch_number,
            'product_name': stock.name,
            'category': stock.category,
            'supplier': stock.supplier.supplier_name if stock.supplier else '',
            'manufacturing_date': b.manufacturing_date,
            'expiry_date': b.expiry_date,
            'quantity_available': b.quantity_available,
            'received_quantity': received_total,
            'storage_location': b.storage_location,
            'assigned_barcode': b.assigned_barcode,
            'status': status,
        })

    # --- Dropdown filters ---
    categories = Batch.objects.select_related('stock').values_list('stock__category', flat=True).distinct()
    suppliers = Supplier.objects.values_list('supplier_name', flat=True).distinct()

    context = {
        'batches': batch_data,
        'search_query': search_query,
        'category_filter': category_filter,
        'supplier_filter': supplier_filter,
        'categories': categories,
        'suppliers': suppliers,
        'generated_on': timezone.now(),
    }

    return render(request, 'reports/batch_report.html', context)

@login_required
def batch_report_excel(request):
    context = get_batch_report_data(request)
    
    # Render Excel content
    response = render(request, "reports/batch_report_export.html", context)
    response["Content-Type"] = "application/vnd.ms-excel"

    # Filename
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Batch_Report_{timestamp}.xls"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Save to media folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        f.write(response.content)

    # Log in database
    ReportFile.objects.create(
        report_type="Batch",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


@login_required
def batch_report_pdf(request):
    context = get_batch_report_data(request)

    # Render PDF content
    html_string = render_to_string("reports/batch_report_export.html", context)
    pdf = HTML(string=html_string).write_pdf()

    # Filename
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Batch_Report_{timestamp}.pdf"

    # Save to media folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        f.write(pdf)

    # Log in database
    ReportFile.objects.create(
        report_type="Batch",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    # Return PDF as download
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

from django.utils import timezone

def get_batch_report_data(request):
    batches = (
        Batch.objects
        .select_related("stock", "stock__product")
        .all()
        .order_by("-created_at")
    )

    report_data = []
    for b in batches:
        product_name = (
            b.stock.product.name
            if b.stock and b.stock.product
            else "-"
        )

        report_data.append({
            "batch_number": b.batch_number,
            "product": product_name,
            "mfg_date": b.manufacturing_date,
            "expiry_date": b.expiry_date,
            "quantity": b.quantity_available,
            "storage": b.storage_location or "-",
            "barcode": b.assigned_barcode or "-",
            "status": b.status,
        })

    return {
        "report_data": report_data,
        "generated_on": timezone.now()
    }


@login_required
def picking_report(request):
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('verification_status', '').strip()

    # Fetch picklists with related items and product info
    picklists = PickList.objects.select_related(
        'order',
        'created_by'
    ).prefetch_related(
        'items',
        'items__order_item',
        'items__order_item__stock__product',  # fetch product from stock
        'items__stock'
    ).order_by('-created_at')

    report_data = []

    for picklist in picklists:
        items = picklist.items.all()

        total_qty = sum(i.order_item.quantity for i in items)
        picked_qty = sum(i.quantity for i in items)

        # Determine dynamic verification status
        if picked_qty == 0:
            status = 'Pending'
        elif picked_qty < total_qty:
            status = 'Partially Picked'
        else:
            status = 'Completed'

        if status_filter and status != status_filter:
            continue

        for item in items:
            # Get product safely
            product = item.order_item.stock.product if item.order_item.stock else None
            product_name = product.name if product else item.order_item.__str__()
            sku = product.sku if product else '-'

            # Determine batch number: prefer PickListItem.batch_no, fallback to stock
            batch_number = item.batch_no or (item.stock.batch_number if item.stock and hasattr(item.stock, 'batch_number') else '-')

            # Search filter
            if search_query and not (
                search_query.lower() in picklist.order.order_no.lower()
                or (product_name and search_query.lower() in product_name.lower())
                or (sku and search_query.lower() in sku.lower())
                or search_query.lower() in str(batch_number).lower()
            ):
                continue

            # Verified By: who picked the item
            verified_by = picklist.created_by.username if picklist.created_by else "-"

            report_data.append({
                'pick_list_no': picklist.id,
                'sales_order_id': picklist.order.order_no,
                'product_name': product_name,
                'product_batch': batch_number,
                'sku': sku,
                'qty_to_pick': item.order_item.quantity,
                'qty_picked': item.quantity,
                'picker_name': picklist.created_by.username if picklist.created_by else '-',
                'picked_date': picklist.created_at.strftime("%Y-%m-%d"),
                'verification_status': status,
                'verified_by': verified_by,
            })

    context = {
        'data': report_data,
        'verification_filter': status_filter,
        'all_statuses': ['Pending', 'Partially Picked', 'Completed'],
        'search_query': search_query,
        'generated_on': timezone.now(),
    }

    return render(request, 'reports/picking_report.html', context)




from django.utils import timezone
#from salesofficer.models import PickList

from django.utils import timezone
#from salesofficer.models import PickList

def get_picking_report_data(request):
    search_query = request.GET.get('search', '').strip().lower()
    status_filter = request.GET.get('verification_status', '').strip()

    # Prefetch related objects for efficiency
    picklists = PickList.objects.select_related(
        'order',
        'created_by',
    ).prefetch_related(
        'items',
        'items__order_item',
        'items__order_item__stock',
        'items__order_item__stock__product',
        'items__stock',
        'deliveryassignment_set__delivery_boy__user',
    ).order_by('-created_at')

    report_data = []

    for picklist in picklists:
        items = picklist.items.all()
        total_qty = sum(i.order_item.quantity for i in items)
        picked_qty = sum(i.quantity for i in items)

        # Determine dynamic status
        if picked_qty == 0:
            status = "Pending"
        elif picked_qty < total_qty:
            status = "Partially Picked"
        else:
            status = "Completed"

        # Apply status filter
        if status_filter and status != status_filter:
            continue

        # Collect verified by users
        verified_by_list = [
            da.delivery_boy.user.username
            for da in picklist.deliveryassignment_set.all()
            if da.delivery_boy and da.delivery_boy.user
        ]
        verified_by = ", ".join(verified_by_list) if verified_by_list else "-"

        for item in items:
            product = item.order_item.stock.product if item.order_item.stock else None
            stock = item.stock or item.order_item.stock

            # Construct product / batch display
            batch_display = stock.batch_number if stock and stock.batch_number else item.batch_no or "-"
            product_batch = f"{product.name if product else '-'} / {batch_display}"

            # Search filter (order no, product name, SKU, batch)
            if search_query:
                if not (
                    search_query in picklist.order.order_no.lower()
                    or (product.name.lower() if product else '').find(search_query) != -1
                    or (product.sku.lower() if product else '').find(search_query) != -1
                    or (batch_display.lower()).find(search_query) != -1
                ):
                    continue

            # Append row
            report_data.append({
                "pick_list_no": picklist.id,
                "sales_order": picklist.order.order_no,
                "product_batch": product_batch,
                "qty_to_pick": item.order_item.quantity,
                "qty_picked": item.quantity,
                "picker": picklist.created_by.username if picklist.created_by else "-",
                "picked_date": picklist.created_at,
                "verification_status": status,
                "verified_by": verified_by,
            })

    return {
        "report_data": report_data,
        "generated_on": timezone.now(),
        "all_statuses": ["Pending", "Partially Picked", "Completed"],
        "verification_filter": status_filter,
        "search_query": search_query,
    }




@login_required
def picking_report_excel(request):
    context = get_picking_report_data(request)

    # Render Excel content
    response = render(request, "reports/picking_report_export.html", context)
    response["Content-Type"] = "application/vnd.ms-excel"

    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Picking_Report_{timestamp}.xls"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Save to media folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        f.write(response.content)

    # Log in database
    ReportFile.objects.create(
        report_type="Picking",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


from django.template.loader import render_to_string
from weasyprint import HTML

@login_required
def picking_report_pdf(request):
    context = get_picking_report_data(request)

    html_string = render_to_string("reports/picking_report_export.html", context)

    output = BytesIO()
    HTML(string=html_string).write_pdf(output)
    output.seek(0)

    # Save in media/reports and DB
    file_name = f"Picking_Report_{context['generated_on'].strftime('%Y%m%d_%H%M')}.pdf"
    save_report(file_name, "Picking", output.read(), request.user)

    # Send as download
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response

# 📦 Packing Report - Web View
from django.db.models import Sum

from django.db.models import Q, Sum
from django.utils import timezone

from django.db.models import Sum, F, Q
from django.utils import timezone
from django.shortcuts import render


def packing_report(request):
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()

    # Prefetch everything needed to avoid N+1 queries
    packings_qs = Packing.objects.select_related(
        'delivery_order', 'packed_by'
    ).prefetch_related(
        'boxes',
        'boxes__items',
        'boxes__items__order_item',
        'boxes__items__order_item__stock__product'
    ).order_by('-packing_date')

    # 🔍 SEARCH FILTER
    if search_query:
        packings_qs = packings_qs.filter(
            Q(id__icontains=search_query) |
            Q(delivery_order__delivery_no__icontains=search_query) |
            Q(boxes__items__order_item__stock__product__name__icontains=search_query) |
            Q(packed_by__username__icontains=search_query)
        ).distinct()

    # 🔽 STATUS FILTER
    if status_filter:
        packings_qs = packings_qs.filter(status=status_filter)

    # Build packing list
    packing_list = []

    for packing in packings_qs:
        boxes = packing.boxes.all()

        box_count = boxes.count()
        total_weight = boxes.aggregate(
            total=Sum('weight')
        )['total'] or 0

        # Efficient total quantity calculation
        total_qty = sum(
            item.order_item.quantity
            for box in boxes
            for item in box.items.all()
        )

        first_item = (
            boxes.first().items.first()
            if boxes.exists() and boxes.first().items.exists()
            else None
        )

        packing_list.append({
            'packing': packing,
            'first_item': first_item,
            'total_qty': total_qty,
            'box_count': box_count,
            'total_weight': total_weight,
        })

    return render(request, 'reports/packing_report.html', {
        'packings': packing_list,
        'search_query': search_query,
        'status_filter': status_filter,
        'generated_on': timezone.now(),
    })







@login_required
def packing_report_excel(request):
    context = get_packing_report_data(request)

    # Render Excel content
    response = render(request, "reports/packing_report_export.html", context)
    response["Content-Type"] = "application/vnd.ms-excel"

    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Packing_Report_{timestamp}.xls"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Save to media folder by year/month
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        f.write(response.content)

    # Log in database
    ReportFile.objects.create(
        report_type="Packing",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


@login_required
def packing_report_pdf(request):
    context = get_packing_report_data(request)

    html_string = render_to_string("reports/packing_report_export.html", context)

    response = HttpResponse(content_type="application/pdf")
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Packing_Report_{timestamp}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    HTML(string=html_string).write_pdf(response)

    # Save PDF to media folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        HTML(string=html_string).write_pdf(f)

    # Log in database
    ReportFile.objects.create(
        report_type="Packing",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


def get_packing_report_data(request):
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()

    packings_qs = Packing.objects.select_related(
        'delivery_order', 'packed_by'
    ).prefetch_related(
        'boxes',
        'boxes__items',
        'boxes__items__order_item',
        'boxes__items__order_item__stock__product'
    ).order_by('-packing_date')

    if search_query:
        packings_qs = packings_qs.filter(
            Q(id__icontains=search_query) |
            Q(delivery_order__delivery_no__icontains=search_query) |
            Q(boxes__items__order_item__stock__product__name__icontains=search_query) |
            Q(packed_by__username__icontains=search_query)
        ).distinct()

    if status_filter:
        packings_qs = packings_qs.filter(status=status_filter)

    report_data = []

    for packing in packings_qs:
        boxes = packing.boxes.all()

        box_count = boxes.count()
        total_weight = boxes.aggregate(total=Sum('weight'))['total'] or 0

        total_qty = sum(
            item.order_item.quantity
            for box in boxes
            for item in box.items.all()
        )

        first_item = (
            boxes.first().items.first()
            if boxes.exists() and boxes.first().items.exists()
            else None
        )

        report_data.append({
            "packing_id": packing.id,
            "delivery_no": packing.delivery_order.delivery_no if packing.delivery_order else "-",
            "product": (
                first_item.order_item.stock.product.name
                if first_item else "-"
            ),
            "total_qty": total_qty,
            "packed_by": packing.packed_by.username if packing.packed_by else "-",
            "packing_date": packing.packing_date,
            "box_count": box_count,
            "total_weight": total_weight,
            "status": packing.status,
        })

    return {
        "report_data": report_data,
        "generated_on": timezone.now(),
        "search_query": search_query,
        "status_filter": status_filter,
    }

@login_required
def dispatch_report(request):
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()

    dispatches = (
        Dispatch.objects
        .select_related('dispatched_by', 'approved_by')
        .prefetch_related(
            'orders',
            Prefetch(
                'orders__picklist_set__deliveryassignment_set',
                queryset=DeliveryAssignment.objects.select_related('delivery_boy'),
                to_attr='delivery_assignments'
            )
        )
        .order_by('-dispatch_date')
    )

    # Search filter
    if search_query:
        dispatches = dispatches.filter(
            Q(orders__order_no__icontains=search_query) |
            Q(carrier_name__icontains=search_query) |
            Q(orders__customer_address__icontains=search_query)
        ).distinct()

    # Delivery status filter
    if status_filter:
        filtered = []
        for dispatch in dispatches:
            for order in dispatch.orders.all():
                for picklist in order.picklist_set.all():
                    assignments = getattr(picklist, 'delivery_assignments', [])
                    if any(a.delivery_status == status_filter for a in assignments):
                        filtered.append(dispatch)
                        break
        dispatches = filtered

    return render(request, 'reports/dispatch_report.html', {
        'dispatches': dispatches,
        'search_query': search_query,
        'status_filter': status_filter,
        'generated_on': timezone.now()
    })

def get_dispatch_report_data(request):
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()

    dispatches = (
        Dispatch.objects
        .select_related('dispatched_by', 'approved_by')
        .prefetch_related(
            'orders',
            Prefetch(
                'orders__picklist_set__deliveryassignment_set',
                queryset=DeliveryAssignment.objects.select_related(
                    'delivery_boy__user'
                ),
                to_attr='delivery_assignments'
            )
        )
        .order_by('-dispatch_date')
    )

    if search_query:
        dispatches = dispatches.filter(
            Q(orders__order_no__icontains=search_query) |
            Q(carrier_name__icontains=search_query) |
            Q(orders__customer_address__icontains=search_query)
        ).distinct()

    report_data = []

    for dispatch in dispatches:
        for order in dispatch.orders.all():
            delivery_statuses = []
            delivery_boys = []

            for picklist in order.picklist_set.all():
                assignments = getattr(picklist, 'delivery_assignments', [])
                for a in assignments:
                    delivery_statuses.append(a.delivery_status)
                    if a.delivery_boy and a.delivery_boy.user:
                        delivery_boys.append(a.delivery_boy.user.username)

            # Decide final status
            if delivery_statuses:
                final_status = "Delivered" if all(
                    s == "Delivered" for s in delivery_statuses
                ) else "In Transit"
            else:
                final_status = "Pending"

            if status_filter and final_status != status_filter:
                continue

            report_data.append({
                "dispatch_id": dispatch.id,
                "order_no": order.order_no,
                "carrier": dispatch.carrier_name or "-",
                "tracking": dispatch.tracking_number or "-",
                "dispatch_date": dispatch.dispatch_date,
                "destination": order.customer_address or "-",
                "status": final_status,
                "delivery_boys": ", ".join(set(delivery_boys)) or "-",
            })

    return {
        "report_data": report_data,
        "search_query": search_query,
        "status_filter": status_filter,
        "generated_on": timezone.now(),
    }


import openpyxl
from django.http import HttpResponse

@login_required
def dispatch_report_excel(request):
    context = get_dispatch_report_data(request)

    # Render Excel content
    response = render(request, "reports/dispatch_report_export.html", context)
    response["Content-Type"] = "application/vnd.ms-excel"

    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Dispatch_Report_{timestamp}.xls"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Save to media folder by year/month
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        f.write(response.content)

    # Log in database
    ReportFile.objects.create(
        report_type="Dispatch",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


@login_required
def dispatch_report_pdf(request):
    context = get_dispatch_report_data(request)

    html_string = render_to_string("reports/dispatch_report_export.html", context)

    response = HttpResponse(content_type="application/pdf")
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Dispatch_Report_{timestamp}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    HTML(string=html_string).write_pdf(response)

    # Save PDF to media folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        HTML(string=html_string).write_pdf(f)

    # Log in database
    ReportFile.objects.create(
        report_type="Dispatch",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response

def delivery_report(request):
    deliveries = DeliveryAssignment.objects.select_related(
        'picklist', 'picklist__order', 'delivery_boy__user'
    ).all()

    # --- Search ---
    search_query = request.GET.get('search', '').strip()
    if search_query:
        deliveries = deliveries.filter(
            Q(id__icontains=search_query) |
            Q(picklist__id__icontains=search_query) |
            Q(delivery_boy__user__username__icontains=search_query) |
            Q(picklist__order__customer_name__icontains=search_query) |
            Q(delivery_status__icontains=search_query)
        )

    # --- Filters ---
    partner_filter = request.GET.get('partner', '').strip()
    status_filter = request.GET.get('status', '').strip()

    if partner_filter:
        deliveries = deliveries.filter(delivery_boy__user__username__iexact=partner_filter)
    if status_filter:
        deliveries = deliveries.filter(delivery_status__iexact=status_filter)

    # --- Prepare partner/status lists for dropdowns ---
    partners = DeliveryAssignment.objects.values_list(
        'delivery_boy__user__username', flat=True
    ).distinct().order_by('delivery_boy__user__username')

    statuses = DeliveryAssignment.objects.values_list(
        'delivery_status', flat=True
    ).distinct().order_by('delivery_status')

    # --- Build table data ---
    data = []
    for d in deliveries:
        order = getattr(d.picklist, 'order', None)
        customer_name = getattr(order, 'customer_name', '-') if order else '-'

        data.append({
            'delivery_id': d.id,
            'dispatch_id': getattr(d.picklist, 'id', '-'),
            'delivery_partner': d.delivery_boy.user.username if d.delivery_boy else 'Admin',
            'customer_name': customer_name,
            'delivery_date': d.assigned_date.strftime("%Y-%m-%d") if d.assigned_date else '-',
            'proof_of_delivery': getattr(d, 'proof_of_delivery', '-') or '-',
            'status': d.delivery_status,
            'remarks': d.remarks or '-'
        })

    context = {
        'deliveries': data,
        'partners': partners,
        'statuses': statuses,
        'search_query': search_query,
        'partner_filter': partner_filter,
        'status_filter': status_filter,
        'generated_on': timezone.now()
    }

    return render(request, 'reports/delivery_report.html', context)
from openpyxl import Workbook
from django.http import HttpResponse



def get_delivery_report_data(request):
    search_query = request.GET.get('search', '').strip()
    partner_filter = request.GET.get('partner', '').strip()
    status_filter = request.GET.get('status', '').strip()

    deliveries = DeliveryAssignment.objects.select_related(
        'picklist',
        'picklist__order',
        'delivery_boy__user'
    ).order_by('-assigned_date')

    # 🔍 Search
    if search_query:
        deliveries = deliveries.filter(
            Q(id__icontains=search_query) |
            Q(picklist__id__icontains=search_query) |
            Q(delivery_boy__user__username__icontains=search_query) |
            Q(picklist__order__customer_name__icontains=search_query) |
            Q(delivery_status__icontains=search_query)
        )

    # 🔽 Filters
    if partner_filter:
        deliveries = deliveries.filter(
            delivery_boy__user__username__iexact=partner_filter
        )
    if status_filter:
        deliveries = deliveries.filter(
            delivery_status__iexact=status_filter
        )

    report_data = []
    for d in deliveries:
        order = getattr(d.picklist, 'order', None)

        report_data.append({
            "delivery_id": d.id,
            "dispatch_id": getattr(d.picklist, 'id', '-'),
            "partner": (
                d.delivery_boy.user.username
                if d.delivery_boy else "Admin"
            ),
            "customer": order.customer_name if order else "-",
            "delivery_date": d.assigned_date,
            "proof": getattr(d, 'proof_of_delivery', '-') or '-',
            "status": d.delivery_status,
            "remarks": d.remarks or '-',
        })

    partners = DeliveryAssignment.objects.values_list(
        'delivery_boy__user__username', flat=True
    ).distinct().order_by('delivery_boy__user__username')

    statuses = DeliveryAssignment.objects.values_list(
        'delivery_status', flat=True
    ).distinct().order_by('delivery_status')

    return {
        "report_data": report_data,
        "partners": partners,
        "statuses": statuses,
        "search_query": search_query,
        "partner_filter": partner_filter,
        "status_filter": status_filter,
        "generated_on": timezone.now(),
    }


@login_required
def delivery_report_excel(request):
    context = get_delivery_report_data(request)

    # Render Excel content
    response = render(request, "reports/delivery_report_export.html", context)
    response["Content-Type"] = "application/vnd.ms-excel"

    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Delivery_Report_{timestamp}.xls"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Save to media folder by year/month
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        f.write(response.content)

    # Log in database
    ReportFile.objects.create(
        report_type="Delivery",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


@login_required
def delivery_report_pdf(request):
    context = get_delivery_report_data(request)

    html_string = render_to_string("reports/delivery_report_export.html", context)

    response = HttpResponse(content_type="application/pdf")
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Delivery_Report_{timestamp}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    HTML(string=html_string).write_pdf(response)

    # Save PDF to media folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        HTML(string=html_string).write_pdf(f)

    # Log in database
    ReportFile.objects.create(
        report_type="Delivery",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from django.template.loader import render_to_string
from weasyprint import HTML






from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

#from warehouse_manager.models import Stock


# -----------------------------------------------
# ✅ 1. View: Stock HTML Report
# -----------------------------------------------
@login_required
def stock_report(request):
    """Display stock list in an HTML table."""
    stocks = Stock.objects.select_related('supplier').all().order_by('-created_at')

    context = {
        'stocks': [{
            'name': s.name,
            'category': s.category,
            'supplier': s.supplier.supplier_name if s.supplier else '-',  # 👈 updated
            'size': s.size or '-',
            'color': s.color or '-',
            'batch_number': s.batch_number or '-',
            'sku': s.sku or '-',
            'unit_price': s.unit_price,
            'quantity': s.quantity,
            'rack': s.rack or '-',
            'zone': s.zone or '-',
            'shelf': s.shelf or '-',
            'expiry_date': s.expiry_date.strftime("%d-%m-%Y") if s.expiry_date else '-',
            'status': 'Active' if s.is_active else 'Inactive',
        } for s in stocks],
        'generated_on': timezone.now(),
    }
    return render(request, 'reports/stock_report.html', context)

def get_stock_report_data(request):
    search_query = request.GET.get('search', '').strip()

    stocks_qs = Stock.objects.select_related('supplier').all().order_by('-created_at')
    
    # 🔍 Search filter
    if search_query:
        stocks_qs = stocks_qs.filter(
            Q(name__icontains=search_query) |
            Q(category__icontains=search_query) |
            Q(supplier__supplier_name__icontains=search_query) |
            Q(batch_number__icontains=search_query) |
            Q(sku__icontains=search_query)
        )

    report_data = []
    for s in stocks_qs:
        report_data.append({
            'name': s.name,
            'category': s.category,
            'supplier': s.supplier.supplier_name if s.supplier else '-',
            'size': s.size or '-',
            'color': s.color or '-',
            'batch_number': s.batch_number or '-',
            'sku': s.sku or '-',
            'unit_price': s.unit_price,
            'quantity': s.quantity,
            'rack': s.rack or '-',
            'zone': s.zone or '-',
            'shelf': s.shelf or '-',
            'expiry_date': s.expiry_date.strftime("%d-%m-%Y") if s.expiry_date else '-',
            'status': 'Active' if s.is_active else 'Inactive',
        })

    return {
        "report_data": report_data,
        "search_query": search_query,
        "generated_on": timezone.now(),
    }

# -----------------------------------------------
# ✅ 2. Export Stock Report to Excel
# -----------------------------------------------
@login_required
def stock_report_excel(request):
    context = get_stock_report_data(request)

    # Render Excel
    response = render(request, "reports/stock_report_export.html", context)
    response["Content-Type"] = "application/vnd.ms-excel"

    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Stock_Report_{timestamp}.xls"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Save to year/month folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        f.write(response.content)

    # Log in DB
    ReportFile.objects.create(
        report_type="Stock",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


@login_required
def stock_report_pdf(request):
    context = get_stock_report_data(request)
    html_string = render_to_string("reports/stock_report_export.html", context)

    response = HttpResponse(content_type="application/pdf")
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"Stock_Report_{timestamp}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Write PDF to response
    HTML(string=html_string).write_pdf(response)

    # Save PDF to year/month folder
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)

    with open(file_path, "wb") as f:
        HTML(string=html_string).write_pdf(f)

    # Log in DB
    ReportFile.objects.create(
        report_type="Stock",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response

from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

#from salesofficer.models import SalesOrder,SalesOrderItem  # adjust app name if needed


# --------------------------------------------------
# Sales Order Reports Views
# --------------------------------------------------

from datetime import datetime
from calendar import monthrange, month_name
from collections import Counter

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string

from openpyxl import Workbook
from weasyprint import HTML

#from salesofficer.models import SalesOrder  # adjust import to your app

# views.py
from datetime import datetime
from calendar import monthrange, month_name
from collections import Counter

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string

#from salesofficer.models import SalesOrder
from openpyxl import Workbook
from weasyprint import HTML
from django.core.mail import EmailMessage

# ------------------------------
def get_filtered_sales_orders(request):
    qs = SalesOrder.objects.all()

    search = request.GET.get('search')
    status = request.GET.get('status')
    month = request.GET.get('month')
    year = request.GET.get('year')

    if search:
        qs = qs.filter(
            Q(order_no__icontains=search) |
            Q(customer_name__icontains=search)
        )

    if status:
        qs = qs.filter(status=status)

    if month:
        qs = qs.filter(order_date__month=month)

    if year:
        qs = qs.filter(order_date__year=year)

    return qs

# ------------------------------
# 1. HTML Report View
# ------------------------------
from django.shortcuts import render
from django.db.models import Sum, Value
from django.db.models.functions import Coalesce
from django.utils import timezone
#from finance.models import RefundInvoice
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Value, DecimalField, Q


from django.shortcuts import render
from django.db.models import Sum, Value, DecimalField, Q
from django.db.models.functions import Coalesce, ExtractMonth
from django.utils import timezone
  # adjust if needed
from django.contrib.auth.decorators import login_required



from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from collections import Counter
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce

#from salesofficer.models import SalesOrder, SalesOrderItem


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from collections import Counter
from datetime import datetime

from calendar import month_name

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, DecimalField
from django.db.models.functions import Coalesce
from collections import Counter
from calendar import month_name

#from salesofficer.models import SalesOrder, SalesOrderItem
#from finance.models import RefundInvoice

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, DecimalField
from django.db.models.functions import Coalesce
from collections import Counter
from calendar import month_name

#from salesofficer.models import SalesOrder, SalesOrderItem
#from finance.models import RefundInvoice

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, DecimalField
from django.db.models.functions import Coalesce
from collections import Counter
from calendar import month_name

#from salesofficer.models import SalesOrder, SalesOrderItem
#from finance.models import RefundInvoice



month_name = {
    1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr',
    5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug',
    9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
}

from django.shortcuts import render
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F
from collections import Counter
from calendar import month_name

#from salesofficer.models import SalesOrder, SalesOrderItem

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Sum
from collections import Counter
from calendar import month_name
from django.db.models import Sum


@login_required
def sales_order_report(request):
    orders_qs = get_filtered_sales_orders(request)

    # Table data
    orders = []
    for o in orders_qs:
        refund_amount = (
            o.refund_invoices.aggregate(
                total=Sum('total_refund_amount')
            )['total'] or 0
        )

        orders.append({
            'order_no': o.order_no,
            'customer_name': o.customer_name,
            'order_date': o.order_date,
            'status_display': o.get_status_display(),
            'total_amount': float(o.total_amount or 0),
            'refund_amount': float(refund_amount),
        })

    # KPIs
    total_orders = orders_qs.count()
    total_sales = orders_qs.aggregate(
        total=Sum('total_amount')
    )['total'] or 0

    total_refund = (
        RefundInvoice.objects
        .filter(order__in=orders_qs)
        .aggregate(total=Sum('total_refund_amount'))
        ['total'] or 0
    )


    # Monthly chart (respect selected year)
    selected_year = request.GET.get('year')
    selected_year = int(selected_year) if selected_year else timezone.now().year

    monthly_totals = [0.0] * 12
    for o in orders_qs:
        if o.order_date.year == selected_year:
            monthly_totals[o.order_date.month - 1] += float(o.total_amount or 0)

    month_labels = [month_name[i] for i in range(1, 13)]

    # Status pie
    status_counter = Counter(
        orders_qs.values_list('status', flat=True)
    )
    status_labels = [
        label for key, label in SalesOrder.STATUS_CHOICES
        if key in status_counter
    ]
    status_counts = [
        status_counter[key] for key, label in SalesOrder.STATUS_CHOICES
        if key in status_counter
    ]

    # Trending products

 

    trending_qs = (
        SalesOrderItem.objects
        .filter(order__in=orders_qs)
        .values('stock__product__name')
        .annotate(total_quantity=Sum('quantity'))
        .order_by('-total_quantity')[:10]
    )

    trending_products = list(trending_qs)

    trending_names = [item['stock__product__name'] for item in trending_products]
    trending_qtys = [item['total_quantity'] for item in trending_products]



    context = {
        'orders': orders,
        'total_orders': total_orders,
        'total_sales': float(total_sales),
        'total_refund': float(total_refund),
        'status_choices': SalesOrder.STATUS_CHOICES,
        'search_query': request.GET.get('search', ''),
        'status_filter': request.GET.get('status', ''),
        'month': request.GET.get('month', ''),
        'year': request.GET.get('year', ''),
        'months': [{'number': i, 'name': month_name[i]} for i in range(1, 13)],
        'years': list(range(2020, timezone.now().year + 1)),
        'month_labels': month_labels,
        'monthly_totals': monthly_totals,
        'status_labels': status_labels,
        'status_counts': status_counts,
        'trending_names': trending_names,
        'trending_qtys': trending_qtys,
        'trending_products': trending_products,


    }

    return render(request, 'reports/salesorder_report.html', context)

















# ------------------------------
# 2. Excel Export
# ------------------------------
@login_required
def sales_order_report_excel(request):
    orders_qs = get_filtered_sales_orders(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Orders"

    # -----------------
    # Orders Sheet
    # -----------------
    headers = ["Order No","Customer Name","Address","Phone","Email","Remarks","Order Date","Status"]
    ws.append(headers)

    for o in orders_qs:
        ws.append([
            o.order_no,
            o.customer_name,
            o.customer_address or '-',
            o.customer_phone or '-',
            o.customer_email or '-',
            o.remarks or '-',
            o.order_date.strftime("%d-%m-%Y %H:%M"),
            o.get_status_display(),
        ])

    # -----------------
    # Trending Products Sheet
    # -----------------
    trending_qs = (
        SalesOrderItem.objects
        .filter(order__in=orders_qs)
        .values('stock__product__name')
        .annotate(total_quantity=Sum('quantity'))
        .order_by('-total_quantity')[:10]
    )

    if trending_qs:
        ws2 = wb.create_sheet(title="Trending Products")
        ws2.append(["Product Name", "Quantity Sold"])
        for item in trending_qs:
            ws2.append([item['stock__product__name'], item['total_quantity']])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"SalesOrders_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ------------------------------
# 3. PDF Export
# ------------------------------


from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from weasyprint import HTML
from collections import Counter
from calendar import month_name
  # your existing filter function

import matplotlib
matplotlib.use('Agg')  # prevents Tkinter GUI errors on server
import matplotlib.pyplot as plt
import io
import base64

# Chart generation function
def generate_chart_image(x_data, y_data, title='Chart', chart_type='bar'):
    plt.figure(figsize=(6,4))

    if chart_type == 'bar':
        plt.bar(x_data, y_data, color='skyblue')
    elif chart_type == 'line':
        plt.plot(x_data, y_data, marker='o', linestyle='-', color='skyblue')
    elif chart_type == 'pie':
        plt.pie(y_data, labels=x_data, autopct='%1.1f%%')
    else:
        raise ValueError(f"Unsupported chart type: {chart_type}")

    plt.title(title)
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format='png')
    plt.close()
    buf.seek(0)

    return base64.b64encode(buf.getvalue()).decode('utf-8')


from collections import Counter
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from calendar import month_name

from weasyprint import HTML


@login_required
def sales_order_report_pdf(request):
    # -----------------------------
    # Fetch filtered orders
    # -----------------------------
    orders_qs = get_filtered_sales_orders(request)
    current_year = timezone.now().year

    # -----------------------------
    # KPI Calculations
    # -----------------------------
    total_orders = orders_qs.count()
    total_sales = orders_qs.aggregate(total=Sum('total_amount'))['total'] or 0

    # Total refunds (fixed field name)
    total_refunds = RefundInvoice.objects.filter(order__in=orders_qs).aggregate(
        total=Sum('total_refund_amount')
    )['total'] or 0

    # Total fulfilled orders
    fulfilled_orders = orders_qs.filter(status='fulfilled').count()

    kpis = {
        'total_orders': total_orders,
        'total_sales': total_sales,
        'total_refunds': total_refunds,
        'fulfilled_orders': fulfilled_orders,
    }

    # -----------------------------
    # Monthly Sales Overview Chart
    # -----------------------------
    monthly_totals = [0] * 12
    for order in orders_qs.filter(order_date__year=current_year):
        monthly_totals[order.order_date.month - 1] += float(order.total_amount)

    month_labels = [month_name[i] for i in range(1, 13)]
    monthly_sales_chart = generate_chart_image(
        month_labels, monthly_totals, title="Monthly Sales Overview", chart_type='bar'
    )

    # -----------------------------
    # Order Status Distribution Chart
    # -----------------------------
    status_counter = Counter(orders_qs.values_list('status', flat=True))
    status_labels = [label for key, label in SalesOrder.STATUS_CHOICES if key in status_counter]
    status_counts = [status_counter[key] for key, _ in SalesOrder.STATUS_CHOICES if key in status_counter]

    status_distribution_chart = generate_chart_image(
        status_labels, status_counts, title="Order Status Distribution", chart_type='pie'
    )

    # -----------------------------
    # Trending Products Chart
    # -----------------------------
    trending_qs = (
        SalesOrderItem.objects
        .filter(order__in=orders_qs)
        .values('stock__product__name')
        .annotate(total_quantity=Sum('quantity'))
        .order_by('-total_quantity')[:10]
    )
    trending_labels = [item['stock__product__name'] for item in trending_qs]
    trending_counts = [item['total_quantity'] for item in trending_qs]

    trending_products_chart = generate_chart_image(
        trending_labels, trending_counts, title="Trending Products", chart_type="bar"
    )

    # -----------------------------
    # Render PDF
    # -----------------------------
    context = {
        'orders': orders_qs,
        'generated_on': timezone.now(),
        'kpis': kpis,
        'monthly_sales_chart': monthly_sales_chart,
        'status_distribution_chart': status_distribution_chart,
        'trending_products': trending_qs,
        'trending_products_chart': trending_products_chart,
    }

    html_string = render_to_string("reports/sales_order_report_export.html", context)
    response = HttpResponse(content_type='application/pdf')
    filename = f"SalesOrders_{timezone.now():%Y%m%d_%H%M%S}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    HTML(string=html_string).write_pdf(response)
    
    return response



# ------------------------------
# 4. Chart Data (JSON)
# ------------------------------
from django.http import JsonResponse
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth
from django.contrib.auth.decorators import login_required


from django.db.models.functions import TruncMonth
from django.db.models import Count, Sum

@login_required
def sales_order_chart_data(request):
    orders_qs = get_filtered_sales_orders(request)

    # KPIs
    total_orders = orders_qs.count()
    total_sales = orders_qs.aggregate(
        total=Sum('total_amount')
    )['total'] or 0

    # Refund
    total_refund = (
        RefundInvoice.objects
        .filter(sales_order__in=orders_qs)
        .aggregate(total=Sum('refund_amount'))
        ['total'] or 0
    )

    # Monthly sales
    monthly_sales = (
        orders_qs
        .annotate(month=TruncMonth('order_date'))
        .values('month')
        .annotate(total=Sum('total_amount'))
        .order_by('month')
    )

    month_labels = [m['month'].strftime('%b %Y') for m in monthly_sales]
    monthly_totals = [float(m['total']) for m in monthly_sales]

    # Status pie
    status_data = orders_qs.values('status').annotate(count=Count('id'))
    status_labels = [s['status'] for s in status_data]
    status_counts = [s['count'] for s in status_data]

    # Trending products
    trending_qs = (
        SalesOrderItem.objects
        .filter(order__in=orders_qs)
        .values('stock__product__name')
        .annotate(total_quantity=Sum('quantity'))
        .order_by('-total_quantity')[:10]
    )

    trending_names = [t['stock__product__name'] for t in trending_qs]
    trending_qtys = [t['total_quantity'] for t in trending_qs]

    # Orders table
    orders = [{
        'order_no': o.order_no,
        'customer_name': o.customer_name,
        'order_date': o.order_date.strftime('%d-%m-%Y'),
        'status_display': o.get_status_display(),
        'total_amount': float(o.total_amount),
    } for o in orders_qs]

    return JsonResponse({
        'orders': orders,
        'month_labels': month_labels,
        'monthly_totals': monthly_totals,
        'status_labels': status_labels,
        'status_counts': status_counts,
        'trending_names': trending_names,
        'trending_qtys': trending_qtys,
        'total_refund': float(total_refund),
    })

# ------------------------------
# 5. Send Email Report to Admin
# ------------------------------
@login_required
def send_sales_report_email(request):
    orders_qs = get_filtered_sales_orders(request)
    context = {'orders': orders_qs}
    html_string = render_to_string("reports/sales_order_report_export.html", context)

    email = EmailMessage(
        subject=f"Sales Order Report - {timezone.now().strftime('%d-%m-%Y')}",
        body=html_string,
        from_email='noreply@example.com',
        to=['admin@example.com'],
    )
    email.content_subtype = 'html'
    email.send()

    return HttpResponse("Email sent successfully!")


from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from decimal import Decimal

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

#from finance.models import SalesInvoice, PurchaseInvoice


# --------------------------------------------------
# ✅ 1. SALES INVOICE REPORT
# --------------------------------------------------
@login_required
def sales_invoice_report(request):
    invoices = SalesInvoice.objects.select_related('order').all().order_by('-invoice_date')

    context = {
        'invoices': invoices,
        'generated_on': timezone.now(),
    }
    return render(request, 'reports/sales_invoice_report.html', context)

# reports/utils.py
from django.utils import timezone
#from finance.models import SalesInvoice

from datetime import datetime
from calendar import monthrange
from django.db.models import Q

from datetime import datetime
from calendar import monthrange
from django.db.models import Q, Sum, Count
from django.utils import timezone
#from salesofficer.models import SalesOrder

from datetime import datetime
from calendar import monthrange
from django.db.models import Q, Sum, Count
from django.utils import timezone
#from salesofficer.models import SalesOrder

from calendar import month_name
from django.utils import timezone

def get_sales_invoice_report_data(request):
    search_query = request.GET.get('search', '').strip()
    month = request.GET.get('month')
    year = request.GET.get('year')

    invoices_qs = SalesInvoice.objects.select_related('order').all().order_by('-invoice_date')

    # Search filter
    if search_query:
        invoices_qs = invoices_qs.filter(
            Q(invoice_no__icontains=search_query) |
            Q(order__order_no__icontains=search_query) |
            Q(order__customer_name__icontains=search_query)
        )

    if year:
        invoices_qs = invoices_qs.filter(invoice_date__year=int(year))
    if month:
        invoices_qs = invoices_qs.filter(invoice_date__month=int(month))

    invoices = []
    for i in invoices_qs:
        invoices.append({
            'invoice_no': i.invoice_no,
            'order_no': i.order.order_no,
            'customer_name': i.order.customer_name,
            'invoice_date': i.invoice_date,
            'subtotal': i.subtotal,
            'vat': i.vat,
            'total': i.total,
            'status': i.status,
            'status_display': i.get_status_display(),
        })

    total_revenue = invoices_qs.aggregate(total=Sum('total'))['total'] or 0
    total_invoices = invoices_qs.count()

    # ✅ Add months and years for dropdowns
    months = [{'number': i, 'name': month_name[i]} for i in range(1, 13)]
    current_year = timezone.now().year
    years = list(range(2020, current_year + 1))

    return {
        'invoices': invoices,
        'search_query': search_query,
        'month': month or '',
        'year': year or '',
        'generated_on': timezone.now(),
        'total_revenue': total_revenue,
        'total_invoices': total_invoices,
        'months': months,
        'years': years,
    }




@login_required
def sales_invoice_report_excel(request):
    context = get_sales_invoice_report_data(request)
    invoices = context['invoices']

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Invoice Report"
    headers = ["Invoice No","Order No","Customer","Date","Subtotal","VAT","Total","Status"]
    ws.append(headers)

    for i in invoices:
        ws.append([
            i['invoice_no'],
            i['order_no'],
            i['customer_name'],
            i['invoice_date'].strftime("%d-%m-%Y"),
            float(i['subtotal']),
            float(i['vat']),
            float(i['total']),
            i['status_display']
        ])

    # Save to response for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"SalesInvoice_Report_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    # Save to MEDIA_ROOT/reports/year/month
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)
    wb.save(file_path)

    # Log in DB
    ReportFile.objects.create(
        report_type="Salesinvoice",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


@login_required
def sales_invoice_report_pdf(request):
    context = get_sales_invoice_report_data(request)
    html_string = render_to_string("reports/salesinvoice_report_export.html", context)

    response = HttpResponse(content_type='application/pdf')
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"SalesInvoice_Report_{timestamp}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Write PDF to response
    HTML(string=html_string).write_pdf(response)

    # Save PDF to MEDIA_ROOT/reports/year/month
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)
    with open(file_path, "wb") as f:
        HTML(string=html_string).write_pdf(f)

    # Log in DB
    ReportFile.objects.create(
        report_type="Salesinvoice",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response

# --------------------------------------------------
# ✅ 2. PURCHASE INVOICE REPORT
# --------------------------------------------------
@login_required
def purchase_invoice_report(request):
    # ---------------------------
    # 1️⃣ Filters
    # ---------------------------
    search_query = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")

    invoices = PurchaseInvoice.objects.all().order_by("-invoice_date")

    if search_query:
        invoices = invoices.filter(
            invoice_no__icontains=search_query
        ) | invoices.filter(
            purchase_order__order_no__icontains=search_query
        ) | invoices.filter(
            supplier__supplier_name__icontains=search_query
        )

    if status_filter:
        invoices = invoices.filter(status=status_filter)

    # ---------------------------
    # 2️⃣ Context
    # ---------------------------
    context = {
        "invoices": invoices,
        "search_query": search_query,
        "status_filter": status_filter,
        "statuses": ["Pending", "Paid", "Cancelled"],
        "generated_on": timezone.now(),
    }

    return render(request, "reports/purchase_invoice_report.html", context)

def get_purchase_invoice_report_data(request):
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()

    invoices_qs = PurchaseInvoice.objects.select_related('supplier', 'purchase_order').all().order_by('-invoice_date')

    if search_query:
        invoices_qs = invoices_qs.filter(
            Q(invoice_no__icontains=search_query) |
            Q(purchase_order__order_no__icontains=search_query) |
            Q(supplier__supplier_name__icontains=search_query)
        )

    if status_filter:
        invoices_qs = invoices_qs.filter(status=status_filter)

    invoices = []
    for i in invoices_qs:
        invoices.append({
            'invoice_no': i.invoice_no,
            'po_no': i.purchase_order.order_no if i.purchase_order else '-',
            'supplier': i.supplier.supplier_name if i.supplier else '-',
            'invoice_date': i.invoice_date,
            'subtotal': i.subtotal,
            'vat': i.vat_amount,
            'total': i.total_amount,
            'status': i.status,
            'status_display': i.get_status_display() if hasattr(i, 'get_status_display') else i.status
        })

    statuses = PurchaseInvoice.objects.values_list('status', flat=True).distinct().order_by('status')

    context = {
        'invoices': invoices,
        'statuses': statuses,
        'search_query': search_query,
        'status_filter': status_filter,
        'generated_on': timezone.now(),
    }
    return context


@login_required
def purchase_invoice_report_excel(request):
    context = get_purchase_invoice_report_data(request)
    invoices = context['invoices']

    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Invoice Report"
    headers = ["Invoice No","PO No","Supplier","Date","Subtotal","VAT","Total","Status"]
    ws.append(headers)

    for i in invoices:
        ws.append([
            i['invoice_no'],
            i['po_no'],
            i['supplier'],
            i['invoice_date'].strftime("%d-%m-%Y"),
            float(i['subtotal']),
            float(i['vat']),
            float(i['total']),
            i['status_display']
        ])

    # Save to response for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"PurchaseInvoice_Report_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    # Save to MEDIA_ROOT/reports/year/month
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)
    wb.save(file_path)

    # Log in DB
    ReportFile.objects.create(
        report_type="Purchaseinvoice",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response


@login_required
def purchase_invoice_report_pdf(request):
    context = get_purchase_invoice_report_data(request)
    html_string = render_to_string("reports/purchaseinvoice_report_export.html", context)

    response = HttpResponse(content_type='application/pdf')
    timestamp = context["generated_on"].strftime("%Y%m%d_%H%M%S")
    filename = f"PurchaseInvoice_Report_{timestamp}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Write PDF to response
    HTML(string=html_string).write_pdf(response)

    # Save PDF to MEDIA_ROOT/reports/year/month
    year = context["generated_on"].strftime("%Y")
    month = context["generated_on"].strftime("%m")
    folder_path = os.path.join(settings.MEDIA_ROOT, "reports", year, month)
    os.makedirs(folder_path, exist_ok=True)
    file_path = os.path.join(folder_path, filename)
    with open(file_path, "wb") as f:
        HTML(string=html_string).write_pdf(f)

    # Log in DB
    ReportFile.objects.create(
        report_type="Purchaseinvoice",
        file=os.path.join("reports", year, month, filename),
        created_by=request.user
    )

    return response
# --------------------------------------------------
# ✅ 3. FINANCE SUMMARY REPORT
# --------------------------------------------------
@login_required
def finance_summary_report(request):
    total_sales = SalesInvoice.objects.filter(status='Paid').aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00')
    total_purchases = PurchaseInvoice.objects.filter(status='Paid').aggregate(Sum('total_amount'))['total_amount__sum'] or Decimal('0.00')
    profit = total_sales - total_purchases

    context = {
        'total_sales': total_sales,
        'total_purchases': total_purchases,
        'profit': profit,
        'generated_on': timezone.now(),
    }
    return render(request, 'reports/finance_summary.html', context)

@login_required
def user_profile_edit(request):
    profile, created = UserProfile.objects.get_or_create(user=request.user)

    if request.method == "POST":
        form = UserProfileEditForm(
            request.POST,
            request.FILES,
            instance=profile,
            user=request.user
        )
        if form.is_valid():
            form.save()

            # Save User model fields
            request.user.first_name = form.cleaned_data.get("first_name", "")
            request.user.last_name = form.cleaned_data.get("last_name", "")
            request.user.email = form.cleaned_data.get("email", "")
            request.user.save()

            return redirect("user_profile")
    else:
        form = UserProfileEditForm(instance=profile, user=request.user)

    return render(request, "user/user_profile_edit.html", {"form": form})


from django.shortcuts import render, redirect, get_object_or_404
from .models import Role
from .forms import RoleFunctionForm

def assign_role_functions(request, role_name):

    # get role object
    role_obj, created = RoleFunction.objects.get_or_create(role=role_name)

    # all functions
    all_functions = Function.objects.all()

    # currently assigned IDs
    assigned_ids = role_obj.functions.values_list("id", flat=True)

    if request.method == "POST":
        selected_ids = request.POST.getlist("functions")  # IDs of checked boxes
        role_obj.functions.set(selected_ids)  # update ManyToMany

        messages.success(request, f"Functions updated for role {role_name}")
        return redirect("role_list")

    context = {
        "role": role_name,
        "all_functions": all_functions,
        "assigned_ids": assigned_ids,
    }
    return render(request, "roles/assign_roles.html", context)


def role_list(request):
    roles = UserProfile.ROLE_CHOICES
    return render(request, "roles/role_list.html", {"roles": roles})




from django.shortcuts import render, get_object_or_404, redirect

from django.http import HttpResponse

# ---------------------------
# CLIENT CRUD OPERATIONS
# ---------------------------




def carrier_list(request):
    carriers = Carrier.objects.all()
    return render(request, "carriers/carrier_list.html", {"carriers": carriers})

def add_carrier(request):
    if request.method == "POST":
        form = CarrierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Carrier added successfully.")
            return redirect("carrier_list")
    else:
        form = CarrierForm()

    return render(request, "carriers/add_carrier.html", {"form": form})


def edit_carrier(request, pk):
    carrier = get_object_or_404(Carrier, pk=pk)
    
    if request.method == "POST":
        form = CarrierForm(request.POST, instance=carrier)
        if form.is_valid():
            form.save()
            messages.success(request, "Carrier updated successfully.")
            return redirect("carrier_list")
    else:
        form = CarrierForm(instance=carrier)

    return render(request, "carriers/edit_carrier.html", {"form": form, "carrier": carrier})

def delete_carrier(request, pk):
    carrier = get_object_or_404(Carrier, pk=pk)
    carrier.delete()
    messages.success(request, "Carrier deleted successfully.")
    return redirect("carrier_list")

@login_required
def delete_all_notifications(request):
    if request.method == "POST":
        request.user.notifications.all().delete()
        messages.success(request, "All notifications deleted successfully.")
    return redirect("admin_notifications")

def custom_password_reset_confirm(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is None:
        messages.error(request, "Invalid reset link.")
        return redirect('forgot_password')

    # Validate token
    if not default_token_generator.check_token(user, token):
        messages.error(request, "Token expired or invalid.")
        return redirect('forgot_password')

    if request.method == 'POST':
        form = SetNewPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            user.set_password(new_password)
            user.save()

            # ✅ RESET 30-MIN SESSION TIMER AFTER PASSWORD RESET
            UserSessionTimer.objects.update_or_create(
                user=user,
                defaults={'login_time': timezone.now()}
            )

            messages.success(
                request,
                "Password updated successfully. Please login."
            )
            return redirect('user_login')

    else:
        form = SetNewPasswordForm()

    return render(
        request,
        'admin_app/reset_password_confirm.html',
        {'form': form}
    )


from django.contrib.auth.decorators import login_required

@login_required
def password_expired(request):
    return render(request, 'admin_app/password_expired.html')
